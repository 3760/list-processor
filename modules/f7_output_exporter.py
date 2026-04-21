"""
F7: 结果输出模块

功能覆盖：
- F7-01: 输出文件生成（多Sheet Excel：处理摘要 + 各名单数据）
- F7-02: 合规检查结果 Sheet（问题数据不参与后续处理，单独输出）
- F7-03: 重复名单结果 Sheet
- F7-04: 结果提示（三种场景，PRD NQ-09）
- F7-05 ~ F7-07: 字典版本记录、输出路径、事务回滚

PRD NQ-09 结果提示规则：
  ✅ 无问题 → "处理完成，结果可直接上传 CEM 系统"（绿色）
  ⚠️ 有警告 → "处理完成，但发现 X 条问题，建议修复后再上传 CEM 系统"（黄色）
  ❌ 失败   → "处理失败：{错误信息}"（红色）
"""

import polars as pl
from pathlib import Path
from typing import Dict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from core.context import ProcessContext
from infra.log_manager import get_logger
from core.base_module import BaseModule

logger = get_logger(__name__)


def export_results(ctx: ProcessContext, output_path: str) -> Dict[str, str]:
    """
    将处理结果输出为 3 个独立 Excel 文件（PRD 附录 B 对齐）。
    
    [20260420-老谈] ISSUE-03+04 重构：
      原：单个多 Sheet xlsx（一线/三方/HW 各一个 Sheet）
      新：3 个独立 xlsx 文件，每个含该名单的处理摘要 + 数据 + 错误 + 去重
    
    输出文件命名（PRD 附录 B + ISSUE-04 原文件名前缀）：
      - 一线：{原文件名}_一线名单_处理结果_{时间戳}.xlsx  （5 个 Sheet）
      - 三方：{原文件名}_三方系统名单_处理结果_{时间戳}.xlsx （1 个 Sheet）
      - HW：  {原文件名}_HW系统名单_处理结果_{时间戳}.xlsx   （2 个 Sheet）
    
    [20260420-老谈] 优化：输出到日期时间子文件夹，便于批次管理
    
    Parameters
    ----------
    ctx : ProcessContext
        包含 dataframes / error_records / module_results / summary
    output_path : str
        输出目录路径
        
    Returns
    -------
    Dict[str, str]
        各名单的输出文件路径映射 {"yixian": "...", "sanfang": "...", "hw": "..."}
    """
    output_dir = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # [FIX] 创建日期时间子文件夹，便于批次管理
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    batch_dir = output_dir / f"批次_{timestamp}"
    batch_dir.mkdir(parents=True, exist_ok=True)
    
    # 更新上下文中的输出路径为批次文件夹
    ctx.output_path = str(batch_dir)
    
    # 提取原始文件名（不含扩展名）作为前缀 [ISSUE-04]
    yixian_input = ctx.get_input_file("yixian") or ""
    original_filename = Path(yixian_input).stem if yixian_input else "客户名单"
    
    # ── 各名单的输出配置 ──────────────────────────────────────
    # [FIX PRD 附录B] Sheet名称对齐PRD：
    # 一线：处理摘要、原始数据（PRD要求）、合规性检查结果、字典校验结果、重复名单结果
    source_config = {
        "yixian": {
            "label": "一线",
            "file_suffix": "一线名单_处理结果",
            "sheets": ["处理摘要", "原始数据", "合规性检查结果", "字典校验结果", "重复名单结果"],
        },
        "sanfang": {
            "label": "三方",
            "file_suffix": "三方系统名单_处理结果",
            "sheets": ["处理摘要", "原始数据"],
        },
        "hw": {
            "label": "HW",
            "file_suffix": "HW系统名单_处理结果",
            "sheets": ["处理摘要", "原始数据"],
        },
    }
    
    output_paths: Dict[str, str] = {}
    
    for source_key, config in source_config.items():
        df = ctx.get_dataframe(source_key)
        if df is None or len(df) == 0:
            logger.info(f"[F7] {config['label']} 无数据，跳过生成")
            continue
        
        # 构建文件名：{原文件名}_{名单类型}_处理结果_{时间戳}.xlsx
        # [FIX] 输出到批次子文件夹
        filename = f"{original_filename}_{config['file_suffix']}_{timestamp}.xlsx"
        file_path = str(batch_dir / filename)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 写入各 Sheet
        _write_single_source_sheets(wb, ctx, source_key, config["label"])
        
        wb.save(file_path)
        output_paths[source_key] = file_path
        logger.info(f"[F7] {config['label']} 结果已输出至：{file_path}")
    
    return output_paths


def _write_single_source_sheets(
    wb, ctx: ProcessContext, source_key: str, label: str
) -> None:
    """
    为单一名单写入所有 Sheet。
    
    [20260420-老谈] ISSUE-03: 从原来的统一多 Sheet 改为按名单拆分
    """
    df = ctx.get_dataframe(source_key)
    
    # 1. 处理摘要 Sheet
    _write_summary_sheet_for_source(wb, ctx, source_key, label, df)

    # 2. 数据 Sheet（仅合规通过的数据）[FIX PRD 附录B] 统一使用"原始数据"
    _write_data_sheet(wb, "原始数据", df)
    
    # 3. 合规性检查结果 Sheet（仅本来源的错误）
    _write_error_records_sheet_for_source(wb, ctx, source_key)
    
    # 4. 字典校验结果 Sheet（如有）—— 仅一线有此 Sheet
    dict_err = ctx.error_records.get("dict_validation")
    if source_key == "yixian" and dict_err is not None and len(dict_err) > 0:
        _write_dict_validation_sheet(wb, dict_err)
    
    # 5. 重复名单结果 Sheet —— 仅一线有此 Sheet
    if source_key == "yixian":
        _write_repeat_records_sheet_for_source(wb, ctx)


# ── 以下为各 Sheet 写入函数（保持原有逻辑，适配单源模式）──────


def _get_display_status(ctx: ProcessContext) -> str:
    """
    根据处理上下文获取显示用的处理状态。
    
    [20260420-老谈] 优化2.1：修复输出文件中处理状态始终为 running 的问题。
    状态判断逻辑：
    - 如果 ctx.status 已经是 completed/failed，直接使用
    - 否则根据错误记录判断
    """
    # 如果状态已明确设置，直接返回
    if ctx.status in ("completed", "failed"):
        return ctx.status
    
    # 否则根据错误记录判断
    total_failed = 0
    for key in ["yixian", "sanfang", "hw"]:
        err = ctx.error_records.get(key)
        if err is not None and len(err) > 0:
            total_failed += len(err)
    
    # 字典校验错误
    dict_err = ctx.error_records.get("dict_validation")
    if dict_err is not None and len(dict_err) > 0:
        total_failed += len(dict_err)
    
    return "completed" if total_failed == 0 else "failed"


def _write_summary_sheet_for_source(
    wb, ctx: ProcessContext, source_key: str, label: str, df: pl.DataFrame
) -> None:
    """[20260420-老谈] ISSUE-16: 增强处理摘要 Sheet - 补充更多统计项"""
    ws = wb.create_sheet(title="处理摘要")
    
    # [20260420-老谈] 优化2.1：使用修复后的状态获取函数
    display_status = _get_display_status(ctx)
    status_display_map = {
        "completed": "已完成",
        "failed": "失败",
        "running": "处理中",
    }
    
    rows = [
        ["统计项", "值"],
        ["名单类型", label],
        ["处理状态", status_display_map.get(display_status, display_status)],
    ]
    
    # [ISSUE-16] 处理耗时
    if ctx.start_time and ctx.end_time:
        try:
            from datetime import datetime
            start = ctx.start_time
            end = ctx.end_time
            if isinstance(start, str):
                start = datetime.fromisoformat(start)
            if isinstance(end, str):
                end = datetime.fromisoformat(end)
            elapsed = (end - start).total_seconds()
            rows.append(["处理耗时(秒)", f"{elapsed:.2f}"])
        except Exception:
            pass
    
    # [FIX PRD F7-07] 处理摘要需包含原始数据量
    # 从 F2 执行前的数据获取原始数量（如果存在）
    original_count = len(df)
    if hasattr(ctx, 'original_row_counts') and source_key in ctx.original_row_counts:
        original_count = ctx.original_row_counts.get(source_key, len(df))

    rows.extend([
        ["开始时间", str(ctx.start_time)],
        ["结束时间", str(ctx.end_time or "")],
        [f"{label} 原始记录数", original_count],  # [FIX] PRD要求原始数据量
        [f"{label} 处理后记录数", len(df)],
    ])
    
    # 模块中文名称映射
    MODULE_NAMES_CN = {
        "F1": "文件加载",
        "F2": "字段合规检查",
        "F3": "跨名单去重",
        "F4": "数据字典上码",
        "F5": "字典值校验",
        "F6": "名单内部去重",
        "F7": "结果输出",
    }
    
    # [ISSUE-16] 各模块执行结果汇总
    for module in ["F1", "F2", "F3", "F4", "F5", "F6"]:
        mod_result = ctx.module_results.get(module, {})
        success = mod_result.get("success", 0)
        fail = mod_result.get("fail", 0)
        skip = mod_result.get("skip", 0)
        if success or fail or skip:
            module_name = MODULE_NAMES_CN.get(module, module)
            rows.append([f"{module_name} 成功数", success])
            rows.append([f"{module_name} 失败数", fail])
            if skip:
                rows.append([f"{module_name} 跳过数", skip])
    
    # [ISSUE-16] 错误类型分布
    err = ctx.error_records.get(source_key)
    if err is not None and len(err) > 0:
        from collections import Counter
        error_types = [row.get("问题类型", "未知") for row in err.iter_rows(named=True)]
        type_counts = Counter(error_types)
        rows.append(["错误类型分布", ""])
        for err_type, count in sorted(type_counts.items()):
            rows.append(["  " + err_type, count])
    
    # [ISSUE-16] 字典信息
    if hasattr(ctx, 'dict_loader') and ctx.dict_loader is not None:
        rows.append(["字典版本", getattr(ctx.dict_loader, 'md5_hash', '未知') or '未知'])
        rows.append(["字典文件", getattr(ctx.dict_loader, 'file_path', '未知') or '未知'])
    
    # 去重信息
    if ctx.dedup_field:
        rows.append(["去重字段", ctx.dedup_field])
        # F6 去重统计 [FIX PRD F6-03] 术语修正：跳过 → 未知
        f6_result = ctx.module_results.get("F6", {})
        if f6_result:
            rows.append(["去重原始数", f6_result.get("success", 0)])
            rows.append(["去重重复数", f6_result.get("fail", 0)])
            rows.append(["去重未知数（空值）", f6_result.get("skip", 0)])  # [FIX] 术语对齐
    
    for row in rows:
        ws.append(row)
    _style_header(ws, num_rows=1)


def _write_error_records_sheet_for_source(
    wb, ctx: ProcessContext, source_key: str
) -> None:
    """
    写入合规性检查结果 Sheet（F2 错误数据）—— 仅本来源的。
    
    [20260420-老谈] ISSUE-06: 从原来的合并所有来源改为按来源独立输出
    """
    err = ctx.error_records.get(source_key)
    if err is None or len(err) == 0:
        return

    ws = wb.create_sheet(title="合规性检查结果")
    ws.append(err.columns)
    for row in err.iter_rows():
        ws.append(list(row))
    _style_header(ws, num_rows=1)


def _apply_generated_column_style(ws, columns: list) -> None:
    """
    为工作表中的新增列添加浅蓝色底色。
    
    [优化] 便于用户快速识别哪些是原始数据列，哪些是系统生成的列。
    """
    # 识别新增列
    generated_cols = [c for c in columns if _is_generated_column(c)]
    
    if not generated_cols:
        return
    
    gen_fill = PatternFill("solid", fgColor="E2EFDA")  # 浅绿色底色
    header_fill = PatternFill("solid", fgColor="D9D9D9")  # 保持表头灰色
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        for cell in row:
            if cell.column_letter:
                # 获取列名
                col_idx = cell.column - 1
                if col_idx < len(columns):
                    col_name = columns[col_idx]
                    if _is_generated_column(col_name):
                        # 表头用浅蓝色，其他行用浅蓝色
                        if row_idx == 1:
                            cell.fill = PatternFill("solid", fgColor="B4C6E7")  # 浅蓝色表头
                            cell.font = Font(bold=True)
                        else:
                            cell.fill = gen_fill


def _write_dict_validation_sheet(wb, dict_err_df: pl.DataFrame) -> None:
    """
    写入字典校验结果 Sheet（F5 错误数据）。
    
    [20260420-老谈] ISSUE-06+17: 新增独立 Sheet（原与 F2 合并导致覆盖）
    """
    ws = wb.create_sheet(title="字典校验结果")
    ws.append(dict_err_df.columns)
    for row in dict_err_df.iter_rows():
        ws.append(list(row))
    _style_header(ws, num_rows=1)
    # [优化] 为新增列添加浅蓝色底色
    _apply_generated_column_style(ws, list(dict_err_df.columns))


def _write_repeat_records_sheet_for_source(wb, ctx: ProcessContext) -> None:
    """
    写入重复名单结果 Sheet（F6 去重标注）。

    [FIX PRD F6-03] 输出格式要求：
      - _行号：原始数据中的行号
      - _重复键值：去重字段的实际值
      - _出现次数：该键值在数据中出现的总次数
      - _重复标记："原始"/"重复"/"未知"（空值行）
    
    [优化] 新增列（_开头列、Code列等）使用浅蓝色底色标注。
    """
    # 尝试从 module_results.F6.rejected 获取
    repeat_data = None
    f6_result = ctx.module_results.get("F6", {})
    if isinstance(f6_result, dict):
        repeat_data = f6_result.get("rejected")

    # 兜底：从 DataFrame 的去重列获取统计信息
    yixian_df = ctx.get_dataframe("yixian")
    has_dedup = (
        yixian_df is not None
        and "内部去重结果" in yixian_df.columns
    )

    # [FIX PRD F6-03] 优先输出完整的重复行详情（含新增列）
    if repeat_data is not None and len(repeat_data) > 0:
        ws = wb.create_sheet(title="重复名单结果")
        columns = None
        if isinstance(repeat_data, pl.DataFrame):
            columns = list(repeat_data.columns)
            ws.append(columns)
            for row in repeat_data.iter_rows():
                ws.append(list(row))
        else:
            # 如果是其他格式（如 list of dicts）
            for item in repeat_data:
                if isinstance(item, dict):
                    if columns is None:
                        columns = list(item.keys())
                        ws.append(columns)
                    ws.append(list(item.values()))
                elif hasattr(item, '__iter__'):
                    row_list = list(item)
                    if columns is None:
                        columns = [f"列{i+1}" for i in range(len(row_list))]
                        ws.append(columns)
                    ws.append(row_list)
        _style_header(ws, num_rows=1)
        # [优化] 为新增列添加浅蓝色底色
        if columns:
            _apply_generated_column_style(ws, columns)
    elif has_dedup:
        # 无 rejected 数据但有去重标记时，写统计摘要
        ws = wb.create_sheet(title="重复名单结果")
        dedup_col = yixian_df["内部去重结果"]
        orig_count = dedup_col.to_list().count("原始") if dedup_col is not None else 0
        # [FIX] 术语修正：跳过 → 未知
        unknown_count = dedup_col.to_list().count("未知") if dedup_col is not None else 0
        dup_count = dedup_col.to_list().count("重复") if dedup_col is not None else 0
        rows = [
            ["统计项", "值"],
            ["原始数", orig_count],
            ["未知数（空值行）", unknown_count],  # [FIX] 术语对齐PRD
            ["重复数", dup_count],
        ]
        for row in rows:
            ws.append(row)
        _style_header(ws, num_rows=1)


# ── [保留] 旧版通用 Sheet 函数（向后兼容）──────────────────────


def _write_summary_sheet(wb, ctx: ProcessContext) -> None:
    """写入处理摘要 Sheet（旧版兼容接口，内部调用新实现）"""
    # 保留此函数以避免外部引用断裂
    df = ctx.get_dataframe("yixian") or pl.DataFrame()
    _write_summary_sheet_for_source(wb, ctx, "yixian", "一线", df)


def _is_generated_column(col_name: str) -> bool:
    """
    判断是否为系统自动生成的列（而非原始数据列）。
    这些列需要用浅蓝色底色标注。
    
    新增列的识别规则：
    - 下划线开头的列：_行号, _重复键值, _出现次数, _重复标记, _来源, _row_num等
    - 特定后缀：_Code（字典上码列：客户来源_Code, 客户等级_Code）
    - 特定前缀：是否已在一线名单, 是否已在三方名单（跨名单去重标注）
    - 内部去重结果
    """
    if col_name.startswith("_"):
        return True
    if col_name.endswith("_Code"):
        return True
    if col_name in ["是否已在一线名单", "是否已在三方名单"]:
        return True
    if col_name == "内部去重结果":
        return True
    return False


def _write_data_sheet(
    wb, sheet_name: str, df: pl.DataFrame
) -> None:
    """
    写入名单数据 Sheet。
    
    [优化] 新增列使用浅蓝色底色（#B4C6E7表头, #E2EFDA数据行），便于与原始数据列区分。
    """
    ws = wb.create_sheet(title=sheet_name)

    # [FIX] 排除内部附加列，只输出原始数据字段
    output_columns = [c for c in df.columns if not c.startswith("_")]
    output_df = df.select(output_columns)

    # 写入表头
    ws.append(output_df.columns)
    
    # 设置表头样式 - 区分原始列和新增列
    header_fill_original = PatternFill("solid", fgColor="D9D9D9")  # 灰色
    header_fill_gen = PatternFill("solid", fgColor="B4C6E7")  # 浅蓝色
    header_font = Font(bold=True)
    for cell in ws[1]:
        col_idx = cell.column - 1
        if col_idx < len(output_columns):
            col_name = output_columns[col_idx]
            if _is_generated_column(col_name):
                cell.fill = header_fill_gen  # 新增列表头用浅蓝色
            else:
                cell.fill = header_fill_original  # 原始列表头用灰色
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    
    # 写入数据行
    for row in output_df.iter_rows():
        ws.append(list(row))
    
    # [优化] 为新增列添加浅绿色底色
    gen_fill = PatternFill("solid", fgColor="E2EFDA")  # 浅绿色
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            if cell.column_letter:
                col_idx = cell.column - 1
                if col_idx < len(output_columns):
                    col_name = output_columns[col_idx]
                    if _is_generated_column(col_name):
                        cell.fill = gen_fill


def _write_error_records_sheet(wb, ctx: ProcessContext) -> None:
    """
    旧版兼容接口：写入合并的合规性检查结果。
    
    [20260420-老谈] ISSUE-03: 保留此函数避免外部引用断裂，
    内部委托给新的单源版本
    """
    # 兼容旧调用方式，写入一线的错误记录
    _write_error_records_sheet_for_source(wb, ctx, "yixian")
    # 同时写入三方和 HW 的
    for key in ["sanfang", "hw"]:
        _write_error_records_sheet_for_source(wb, ctx, key)


def _write_repeat_records_sheet(wb, ctx: ProcessContext) -> None:
    """旧版兼容接口，内部委托给新实现"""
    _write_repeat_records_sheet_for_source(wb, ctx)


def _style_header(ws, num_rows: int = 1) -> None:
    """设置表头样式（灰色背景加粗）"""
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    for row in ws.iter_rows(max_row=num_rows):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left")


def build_result_message(ctx: ProcessContext) -> tuple[str, str]:
    """
    根据处理上下文生成结果提示信息（PRD NQ-09）。
    
    [20260420-老谈] ISSUE-06 修正：
      原逻辑只遍历 error_records["yixian/sanfang/hw"]，
      现需额外统计 dict_validation（F5 独立 key）的错误数

    Returns
    -------
    tuple[str, str]
        (message, level)
        - level: "success" | "warning" | "error"
    """
    # 统计失败总数（error_records + dict_validation）
    total_failed = 0
    
    # 各来源的错误记录
    for key in ["yixian", "sanfang", "hw"]:
        err = ctx.error_records.get(key)
        if err is not None and len(err) > 0:
            total_failed += len(err)
    
    # [20260420-老谈] ISSUE-06: 补充字典校验错误统计（F5 独立 key）
    dict_err = ctx.error_records.get("dict_validation")
    if dict_err is not None and len(dict_err) > 0:
        total_failed += len(dict_err)

    if ctx.status == "failed":
        # ❌ 失败
        last_error = ""
        for mod_result in ctx.module_results.values():
            if isinstance(mod_result, dict) and "message" in mod_result:
                last_error = mod_result["message"]
                break
        message = f"处理失败：{last_error}"
        level = "error"

    elif total_failed > 0:
        # ⚠️ 有警告
        message = (
            f"处理完成，但发现 {total_failed} 条问题，"
            "建议修复后再上传 CEM 系统。[查看详情]"
        )
        level = "warning"

    else:
        # ✅ 无问题
        message = "处理完成，结果可直接上传 CEM 系统"
        level = "success"

    logger.info(f"[F7] 结果提示（{level}）：{message}")
    return message, level


class OutputExporterModule(BaseModule):
    """
    F7 结果输出模块（BaseModule 封装）。

    负责将处理结果输出为多 Sheet Excel 文件，
    包含处理摘要、各名单数据、合规检查结果、重复名单结果。
    """

    def get_module_name(self) -> str:
        return "F7"

    def validate_input(self, context: ProcessContext) -> tuple[bool, str]:
        """校验：ProcessContext 中应有至少一份 DataFrame"""
        has_data = any(
            context.get_dataframe(key) is not None
            for key in ["yixian", "sanfang", "hw"]
        )
        if not has_data:
            return False, "没有可输出的数据，请先执行文件加载"
        return True, ""

    def execute(self, context: ProcessContext) -> ProcessContext:
        """执行结果输出（[20260420-老谈] ISSUE-03+04: 拆为3个独立文件）"""
        from modules.f7_output_exporter import export_results

        # 确定输出目录 [ISSUE-03: output_path 改为目录语义]
        output_dir = getattr(context, 'output_path', None)
        if not output_dir:
            # 默认输出到输入文件所在目录
            yixian_path = context.get_input_file("yixian")
            if yixian_path:
                import os
                output_dir = os.path.dirname(yixian_path)
            else:
                output_dir = os.getcwd()

        # export_results 现在返回 Dict[source_key -> file_path]
        result_paths = export_results(ctx=context, output_path=output_dir)
        
        # 将结果存入 context：主路径指向一线文件（兼容旧逻辑）
        if result_paths:
            context.output_path = result_paths.get("yixian", list(result_paths.values())[0])
            context._all_output_paths = result_paths  # 新增：所有输出文件映射

        # 记录结果消息
        message, level = build_result_message(context)
        context.record_module_result(
            module="F7",
            success_count=len(result_paths),
            fail_count=0,
            message=f"[F7] {message}",
        )
        logger.info(f"[F7] 结果已输出至 {len(result_paths)} 个文件")
        return context

    def get_progress_weight(self) -> int:
        return 10
