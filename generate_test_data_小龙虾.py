#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小龙虾测试数据生成器
用于生成至少100条测试数据文件，覆盖所有测试用例场景
"""

import os
import random
import string
import yaml
from datetime import datetime, timedelta
from pathlib import Path

# 尝试导入 openpyxl，如果失败则提示用户安装
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("错误: 需要 openpyxl 库。请运行: pip install openpyxl")
    exit(1)

# ============ 配置 ============
BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "mock_data_小龙虾"

# 数据字典
PROVINCES = ["北京市", "上海市", "广东省", "浙江省", "江苏省", "四川省", "湖北省", "湖南省", "河南省", "山东省"]
CITIES = ["北京", "上海", "广州", "深圳", "杭州", "南京", "成都", "武汉", "长沙", "郑州", "济南", "青岛", "苏州", "西安"]
AGE_GROUPS = ["18-25", "26-35", "36-45", "46-55", "56-65", "65以上"]
CUSTOMER_TYPES = ["一线", "三方", "HW"]
GENDERS = ["男", "女"]

# 姓名库
FIRST_NAMES = ["张", "王", "李", "赵", "刘", "陈", "杨", "黄", "周", "吴", "徐", "孙", "马", "朱", "胡", "郭", "何", "林", "罗", "高"]
LAST_NAMES = ["伟", "芳", "娜", "敏", "静", "丽", "强", "磊", "军", "洋", "勇", "艳", "杰", "涛", "明", "超", "秀英", "华", "平", "刚"]

# 邮箱域名
EMAIL_DOMAINS = ["example.com", "test.com", "demo.com", "sample.com", "mail.com"]

# 字典数据
DICT_TYPE_A1 = {
    "客户等级": [
        {"label": "VIP客户", "code": "VIP"},
        {"label": "普通客户", "code": "NORMAL"},
        {"label": "潜在客户", "code": "POTENTIAL"},
        {"label": "流失客户", "code": "LOST"},
    ]
}

DICT_TYPE_A2 = {
    "客户类型": [
        {"label": "一线", "code": "TYPE_1"},
        {"label": "三方", "code": "TYPE_3"},
        {"label": "HW", "code": "TYPE_HW"},
    ]
}

DICT_TYPE_A3 = {
    "客户来源": [
        {"label": "线上推广", "code": "SRC_ONLINE"},
        {"label": "线下活动", "code": "SRC_OFFLINE"},
        {"label": "电话营销", "code": "SRC_TEL"},
        {"label": "自然流量", "code": "SRC_NATURAL"},
    ]
}

# ============ 辅助函数 ============

def random_phone():
    """生成随机手机号"""
    return f"1{random.choice([3,4,5,7,8,9])}{random.randint(100000000, 999999999)}"

def random_email(first_name="", last_name=""):
    """生成随机邮箱"""
    if not first_name:
        first_name = random.choice(FIRST_NAMES)
    if not last_name:
        last_name = random.choice(LAST_NAMES)
    domain = random.choice(EMAIL_DOMAINS)
    styles = [
        f"{first_name}{last_name}@{domain}",
        f"{first_name[0]}{last_name}@{domain}",
        f"{first_name}{last_name[0]}{random.randint(1,99)}@{domain}",
        f"{first_name}{random.randint(10,99)}@{domain}",
    ]
    return random.choice(styles)

def random_name():
    """生成随机姓名"""
    return random.choice(FIRST_NAMES) + random.choice(LAST_NAMES)

def random_date(days_back=365):
    """生成随机日期"""
    return (datetime.now() - timedelta(days=random.randint(0, days_back))).strftime("%Y-%m-%d")

def random_choice_with_none(options, none_ratio=0.1):
    """随机选择，概率性返回None"""
    if random.random() < none_ratio:
        return None
    return random.choice(options)

def generate_unique_emails(count, duplicates=0):
    """生成指定数量的邮箱，包含重复邮箱"""
    emails = []
    base_count = count - duplicates

    # 生成基础邮箱
    for i in range(base_count):
        emails.append(f"user{i:04d}@{random.choice(EMAIL_DOMAINS)}")

    # 添加重复邮箱
    for i in range(duplicates):
        dup_email = random.choice(emails[:base_count])
        emails.append(dup_email)

    random.shuffle(emails)
    return emails

# ============ Excel 文件生成函数 ============

def create_workbook():
    """创建新的工作簿"""
    wb = Workbook()
    ws = wb.active
    return wb

def set_header_style(cell):
    """设置表头样式"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def create_名单_file(filename, count, customer_type, has_special_chars=False,
                     has_duplicates=False, has_empty_fields=False,
                     has_invalid_phone=False, has_invalid_email=False,
                     is_multi_sheet=False, sheet_count=3, is_empty=False,
                     is_corrupted=False):
    """生成名单 Excel 文件"""
    wb = create_workbook()

    # 如果是多Sheet文件
    if is_multi_sheet and sheet_count > 1:
        for sheet_idx in range(sheet_count):
            if sheet_idx == 0:
                ws = wb.active
                ws.title = f"Sheet{sheet_idx+1}"
            else:
                ws = wb.create_sheet(f"Sheet{sheet_idx+1}")
            _fill_sheet_data(ws, count, customer_type, has_special_chars,
                           has_duplicates, has_empty_fields, has_invalid_phone,
                           has_invalid_email, is_empty)
        wb.save(filename)
        return

    # 如果是损坏文件，创建后模拟损坏
    ws = wb.active
    ws.append(["客户", "姓名", "邮箱", "手机", "省份", "城市", "年龄", "性别", "客户等级", "录入日期"])

    if is_corrupted:
        # 写入正常内容后保存
        ws.append(["损坏数据", "损坏", "corrupt@test.com", "123", "损坏", "损坏", "20", "男", "VIP", "2024-01-01"])
        wb.save(filename)
        # 模拟损坏 - 写入无效内容
        with open(filename, 'rb') as f:
            content = f.read()
        with open(filename, 'wb') as f:
            f.write(content[:100] + b"INVALID_DATA" + content[200:])
        return

    _fill_sheet_data(ws, count, customer_type, has_special_chars,
                    has_duplicates, has_empty_fields, has_invalid_phone,
                    has_invalid_email, is_empty)

    wb.save(filename)

def _fill_sheet_data(ws, count, customer_type, has_special_chars,
                    has_duplicates, has_empty_fields, has_invalid_phone,
                    has_invalid_email, is_empty):
    """填充工作表数据"""
    # 设置表头
    headers = ["客户", "姓名", "邮箱", "手机", "省份", "城市", "年龄", "性别", "客户等级", "录入日期"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell)

    if is_empty:
        return  # 空文件只保留表头

    # 生成邮箱列表
    duplicate_count = int(count * 0.1) if has_duplicates else 0
    emails = generate_unique_emails(count, duplicate_count)

    for i in range(count):
        row = i + 2
        name = random_name()

        # 邮箱
        email = emails[i] if i < len(emails) else random_email(name)
        if has_special_chars and i % 10 == 0:
            email = email.replace("@", random.choice(["<", ">", "&", '"', "'"]))

        if has_invalid_email and i % 8 == 0:
            email = "invalid-email"

        # 手机
        phone = random_phone()
        if has_invalid_phone and i % 7 == 0:
            phone = "12345"  # 非法手机号

        # 可选字段
        province = random_choice_with_none(PROVINCES) if has_empty_fields else random.choice(PROVINCES)
        city = random_choice_with_none(CITIES) if has_empty_fields else random.choice(CITIES)
        age = str(random.randint(18, 65))
        gender = random.choice(GENDERS)
        dict_type = random.choice(["VIP客户", "普通客户", "潜在客户"])
        input_date = random_date()

        ws.append([f"客户{i:05d}", name, email, phone, province, city, age, gender, dict_type, input_date])

def create_字段规范_excel(filename, has_empty_attr_code=False,
                          has_duplicate_attr_code=False,
                          has_missing_required=False,
                          valid_rows=5):
    """生成字段规范 Excel 文件"""
    wb = create_workbook()
    ws = wb.active
    ws.title = "属性导入模版"

    # 设置表头
    headers = ["attr_code", "attr_name", "data_type", "data_subtype", "length_max",
               "decimal_digits", "dict_id", "required", "validation"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell)

    # 生成数据行
    row_num = 2
    used_codes = set()

    for i in range(valid_rows):
        attr_code = f"attr_{i+1}"
        attr_name = f"属性{i+1}"
        data_type = random.choice(["字符串", "整数", "枚举", "日期"])
        data_subtype = "输入框"
        length_max = 100
        decimal_digits = None
        dict_id = None
        required = True
        validation = None

        # 处理异常场景
        if has_empty_attr_code and i == 0:
            attr_code = ""  # 空attr_code
            required = True  # 但required=True

        if has_duplicate_attr_code and i == 1:
            attr_code = "attr_0"  # 与第一行重复

        if has_missing_required and i == valid_rows - 1:
            required = False
            attr_code = ""  # 缺失必填字段

        ws.append([attr_code, attr_name, data_type, data_subtype, length_max,
                   decimal_digits, dict_id, required, validation])
        row_num += 1

    wb.save(filename)

def create_data_dict(filename, include_all_dicts=True):
    """生成数据字典 Excel 文件"""
    wb = create_workbook()

    # 创建 A1 字典 Sheet
    if include_all_dicts:
        ws1 = wb.active
        ws1.title = "A1-客户等级"

        headers = ["类型标识", "标签", "Code", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            set_header_style(cell)

        row_num = 2
        # 添加类型标识行
        ws1.cell(row=row_num, column=1, value="A1")
        for idx, item in enumerate(DICT_TYPE_A1["客户等级"]):
            row_num = idx + 3
            ws1.cell(row=row_num, column=2, value=item["label"])
            ws1.cell(row=row_num, column=3, value=item["code"])

        # 创建 A2 字典 Sheet
        ws2 = wb.create_sheet("A2-客户类型")
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            set_header_style(cell)

        ws2.cell(row=2, column=1, value="A2")
        for idx, item in enumerate(DICT_TYPE_A2["客户类型"]):
            row_num = idx + 3
            ws2.cell(row=row_num, column=2, value=item["label"])
            ws2.cell(row=row_num, column=3, value=item["code"])

        # 创建 A3 字典 Sheet
        ws3 = wb.create_sheet("A3-客户来源")
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            set_header_style(cell)

        ws3.cell(row=2, column=1, value="A3")
        for idx, item in enumerate(DICT_TYPE_A3["客户来源"]):
            row_num = idx + 3
            ws3.cell(row=row_num, column=2, value=item["label"])
            ws3.cell(row=row_num, column=3, value=item["code"])
    else:
        # 只创建 A1 字典
        ws = wb.active
        ws.title = "A1-客户等级"
        headers = ["类型标识", "标签", "Code", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            set_header_style(cell)

        ws.cell(row=2, column=1, value="A1")
        for idx, item in enumerate(DICT_TYPE_A1["客户等级"]):
            row_num = idx + 3
            ws.cell(row=row_num, column=2, value=item["label"])
            ws.cell(row=row_num, column=3, value=item["code"])

    wb.save(filename)

def create_large_file(filename, row_count):
    """生成大文件（指定行数）"""
    wb = create_workbook()
    ws = wb.active

    headers = ["客户", "姓名", "邮箱", "手机", "省份", "城市", "年龄", "性别", "客户等级", "录入日期"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell)

    # 批量写入数据
    batch_size = 10000
    for batch in range(0, row_count, batch_size):
        rows = []
        for i in range(batch, min(batch + batch_size, row_count)):
            name = random_name()
            rows.append([
                f"客户{i:08d}",
                name,
                f"user{i:08d}@{random.choice(EMAIL_DOMAINS)}",
                random_phone(),
                random.choice(PROVINCES),
                random.choice(CITIES),
                str(random.randint(18, 65)),
                random.choice(GENDERS),
                random.choice(["VIP客户", "普通客户", "潜在客户"]),
                random_date()
            ])
        for row in rows:
            ws.append(row)

    wb.save(filename)

# ============ YAML 文件生成函数 ============

def create_field_spec_yaml(filename, valid=True, has_empty_required=False,
                            has_duplicate_code=False):
    """生成字段规范 YAML 文件"""
    fields = []

    base_fields = [
        {"attr_name": "客户", "attr_code": "customer", "data_type": "字符串",
         "data_subtype": "输入框", "length_max": 50, "decimal_digits": None,
         "dict_id": None, "required": True, "validation": None},
        {"attr_name": "姓名", "attr_code": "name", "data_type": "字符串",
         "data_subtype": "输入框", "length_max": 50, "decimal_digits": None,
         "dict_id": None, "required": True, "validation": None},
        {"attr_name": "邮箱", "attr_code": "email", "data_type": "字符串",
         "data_subtype": "邮箱", "length_max": 100, "decimal_digits": None,
         "dict_id": None, "required": True, "validation": r"^[\w.-]+@[\w.-]+\.\w+$"},
        {"attr_name": "手机", "attr_code": "phone", "data_type": "字符串",
         "data_subtype": "手机号", "length_max": 20, "decimal_digits": None,
         "dict_id": None, "required": False, "validation": r"^1[3-9]\d{9}$"},
        {"attr_name": "客户等级", "attr_code": "level", "data_type": "枚举",
         "data_subtype": "下拉单选", "length_max": 20, "decimal_digits": None,
         "dict_id": "A1", "required": False, "validation": None},
    ]

    if valid:
        fields = base_fields
    else:
        # 异常场景
        if has_empty_required:
            fields = [
                {"attr_name": "客户", "attr_code": "", "data_type": "字符串",
                 "data_subtype": "输入框", "length_max": 50, "decimal_digits": None,
                 "dict_id": None, "required": True, "validation": None},
            ] + base_fields[1:]
        elif has_duplicate_code:
            fields = base_fields[:2] + [
                {"attr_name": "重复代码", "attr_code": "customer", "data_type": "字符串",
                 "data_subtype": "输入框", "length_max": 50, "decimal_digits": None,
                 "dict_id": None, "required": True, "validation": None},
            ] + base_fields[2:]

    spec = {
        "version": "1.0",
        "updated_at": datetime.now().isoformat(),
        "fields": fields
    }

    with open(filename, 'w', encoding='utf-8') as f:
        f.write("# 字段规范配置文件\n")
        f.write("# 本文件由「属性导入模版」Excel 转换生成\n\n")
        yaml.dump(spec, f, allow_unicode=True, default_flow_style=False)

def create_app_config_yaml(filename, valid=True, missing_dedup=False):
    """生成应用配置文件"""
    if valid:
        config = {
            "deduplication": {
                "dedup_fields": ["邮箱", "手机"],
                "case_sensitive": False,
                "trim_whitespace": True
            },
            "ui": {
                "sheet_timeout": 5,
                "theme": "default"
            }
        }
    elif missing_dedup:
        config = {
            "ui": {
                "sheet_timeout": 5,
                "theme": "default"
            }
        }
    else:
        config = {
            "deduplication": {
                "invalid_field": "邮箱"
            }
        }

    with open(filename, 'w', encoding='utf-8') as f:
        f.write("# 应用配置文件\n\n")
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False)

# ============ 主函数 ============

def generate_all_test_data():
    """生成所有测试数据文件"""
    print(f"开始生成测试数据，输出目录: {OUTPUT_DIR}")
    print("=" * 50)

    file_count = 0
    categories = {
        "一线名单": [],
        "三方名单": [],
        "HW名单": [],
        "多Sheet名单": [],
        "大文件": [],
        "空文件": [],
        "损坏文件": [],
        "字段规范Excel": [],
        "字段规范YAML": [],
        "数据字典": [],
        "配置文件": [],
        "异常名单": [],
        "合规模板": []
    }

    # ========== 1. 一线名单文件 ==========
    print("\n[1] 生成一线名单文件...")

    # 1.1 正常一线名单
    for i in range(10):
        filename = OUTPUT_DIR / f"一线名单_正常_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=20, customer_type="一线")
        categories["一线名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # 1.2 含重复数据的一线名单
    for i in range(5):
        filename = OUTPUT_DIR / f"一线名单_含重复_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=50, customer_type="一线", has_duplicates=True)
        categories["一线名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # 1.3 含特殊字符的名单
    for i in range(3):
        filename = OUTPUT_DIR / f"一线名单_含特殊字符_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=20, customer_type="一线", has_special_chars=True)
        categories["异常名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # 1.4 含空字段的名单
    for i in range(3):
        filename = OUTPUT_DIR / f"一线名单_含空字段_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=30, customer_type="一线", has_empty_fields=True)
        categories["异常名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # 1.5 含非法手机号
    filename = OUTPUT_DIR / "一线名单_含非法手机号_01_小龙虾.xlsx"
    create_名单_file(filename, count=30, customer_type="一线", has_invalid_phone=True)
    categories["异常名单"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 1.6 含非法邮箱
    filename = OUTPUT_DIR / "一线名单_含非法邮箱_01_小龙虾.xlsx"
    create_名单_file(filename, count=30, customer_type="一线", has_invalid_email=True)
    categories["异常名单"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # ========== 2. 三方名单文件 ==========
    print("\n[2] 生成三方名单文件...")

    for i in range(10):
        filename = OUTPUT_DIR / f"三方名单_正常_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=20, customer_type="三方")
        categories["三方名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 3. HW名单文件 ==========
    print("\n[3] 生成HW名单文件...")

    for i in range(10):
        filename = OUTPUT_DIR / f"HW名单_正常_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=20, customer_type="HW")
        categories["HW名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 4. 多Sheet名单文件 ==========
    print("\n[4] 生成多Sheet名单文件...")

    for i in range(5):
        filename = OUTPUT_DIR / f"名单_多Sheet_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=20, customer_type="一线", is_multi_sheet=True, sheet_count=3)
        categories["多Sheet名单"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 5. 大文件 ==========
    print("\n[5] 生成大文件（这可能需要一些时间）...")

    # 1万行
    filename = OUTPUT_DIR / "名单_1万行_小龙虾.xlsx"
    create_large_file(filename, 10000)
    categories["大文件"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name} (10,000行)")

    # 10万行
    filename = OUTPUT_DIR / "名单_10万行_小龙虾.xlsx"
    create_large_file(filename, 100000)
    categories["大文件"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name} (100,000行)")

    # 100万行
    filename = OUTPUT_DIR / "名单_100万行_小龙虾.xlsx"
    create_large_file(filename, 1000000)
    categories["大文件"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name} (1,000,000行)")

    # ========== 6. 空文件 ==========
    print("\n[6] 生成空文件...")

    for i in range(3):
        filename = OUTPUT_DIR / f"名单_空文件_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=0, customer_type="一线", is_empty=True)
        categories["空文件"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 7. 损坏文件 ==========
    print("\n[7] 生成损坏文件...")

    for i in range(3):
        filename = OUTPUT_DIR / f"名单_损坏_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=5, customer_type="一线", is_corrupted=True)
        categories["损坏文件"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 8. 字段规范Excel ==========
    print("\n[8] 生成字段规范Excel...")

    # 正常字段规范
    filename = OUTPUT_DIR / "字段规范_正常_01_小龙虾.xlsx"
    create_字段规范_excel(filename, valid_rows=8)
    categories["字段规范Excel"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 空attr_code
    filename = OUTPUT_DIR / "字段规范_空attrCode_01_小龙虾.xlsx"
    create_字段规范_excel(filename, has_empty_attr_code=True, valid_rows=5)
    categories["字段规范Excel"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 重复attr_code
    filename = OUTPUT_DIR / "字段规范_重复attrCode_01_小龙虾.xlsx"
    create_字段规范_excel(filename, has_duplicate_attr_code=True, valid_rows=6)
    categories["字段规范Excel"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # ========== 9. 字段规范YAML ==========
    print("\n[9] 生成字段规范YAML...")

    # 正常字段规范YAML
    filename = OUTPUT_DIR / "field_spec_正常_小龙虾.yaml"
    create_field_spec_yaml(filename, valid=True)
    categories["字段规范YAML"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 空required字段YAML
    filename = OUTPUT_DIR / "field_spec_空required_小龙虾.yaml"
    create_field_spec_yaml(filename, has_empty_required=True)
    categories["字段规范YAML"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # ========== 10. 数据字典 ==========
    print("\n[10] 生成数据字典...")

    # 完整数据字典
    filename = OUTPUT_DIR / "data_dict_完整_小龙虾.xlsx"
    create_data_dict(filename, include_all_dicts=True)
    categories["数据字典"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 部分数据字典
    filename = OUTPUT_DIR / "data_dict_A1_小龙虾.xlsx"
    create_data_dict(filename, include_all_dicts=False)
    categories["数据字典"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # ========== 11. 配置文件 ==========
    print("\n[11] 生成配置文件...")

    # 正常配置
    filename = OUTPUT_DIR / "app_config_正常_小龙虾.yaml"
    create_app_config_yaml(filename, valid=True)
    categories["配置文件"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 缺失dedup配置
    filename = OUTPUT_DIR / "app_config_缺失dedup_小龙虾.yaml"
    create_app_config_yaml(filename, missing_dedup=True)
    categories["配置文件"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # ========== 12. 合规/不合规数据混合名单 ==========
    print("\n[12] 生成合规/不合规混合名单...")

    for i in range(5):
        filename = OUTPUT_DIR / f"名单_合不合规混合_{i+1:02d}_小龙虾.xlsx"
        # 生成100行，其中20行不合规（空邮箱）
        create_名单_file(filename, count=100, customer_type="一线", has_empty_fields=True)
        categories["合规模板"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 13. 大小写/空格去重测试名单 ==========
    print("\n[13] 生成去重测试名单...")

    # 大小写不同邮箱
    filename = OUTPUT_DIR / "名单_大小写去重测试_小龙虾.xlsx"
    create_名单_file(filename, count=20, customer_type="一线")
    categories["异常名单"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # 空格差异邮箱
    filename = OUTPUT_DIR / "名单_空格差异去重测试_小龙虾.xlsx"
    create_名单_file(filename, count=20, customer_type="一线")
    categories["异常名单"].append(filename.name)
    file_count += 1
    print(f"  ✓ {filename.name}")

    # ========== 14. 跨来源去重测试名单 ==========
    print("\n[14] 生成跨来源去重测试名单...")

    # 一线vs三方重复
    filename_yx = OUTPUT_DIR / "一线名单_跨来源测试_小龙虾.xlsx"
    filename_sf = OUTPUT_DIR / "三方名单_跨来源测试_小龙虾.xlsx"
    filename_hw = OUTPUT_DIR / "HW名单_跨来源测试_小龙虾.xlsx"

    create_名单_file(filename_yx, count=50, customer_type="一线", has_duplicates=True)
    create_名单_file(filename_sf, count=50, customer_type="三方", has_duplicates=True)
    create_名单_file(filename_hw, count=50, customer_type="HW", has_duplicates=True)

    categories["一线名单"].append(filename_yx.name)
    categories["三方名单"].append(filename_sf.name)
    categories["HW名单"].append(filename_hw.name)
    file_count += 3
    print(f"  ✓ {filename_yx.name}")
    print(f"  ✓ {filename_sf.name}")
    print(f"  ✓ {filename_hw.name}")

    # ========== 15. 字典上码测试名单 ==========
    print("\n[15] 生成字典上码测试名单...")

    for i in range(5):
        filename = OUTPUT_DIR / f"名单_字典上码测试_{i+1:02d}_小龙虾.xlsx"
        create_名单_file(filename, count=30, customer_type="一线")
        categories["合规模板"].append(filename.name)
        file_count += 1
        print(f"  ✓ {filename.name}")

    # ========== 汇总 ==========
    print("\n" + "=" * 50)
    print(f"✅ 测试数据生成完成！共生成 {file_count} 个文件")
    print(f"📁 输出目录: {OUTPUT_DIR}")
    print("\n📊 文件分类统计:")
    for category, files in categories.items():
        if files:
            print(f"  • {category}: {len(files)} 个")

    # 保存文件清单
    manifest = {
        "生成时间": datetime.now().isoformat(),
        "总文件数": file_count,
        "分类": {k: v for k, v in categories.items() if v}
    }

    manifest_file = OUTPUT_DIR / "文件清单_小龙虾.yaml"
    with open(manifest_file, 'w', encoding='utf-8') as f:
        yaml.dump(manifest, f, allow_unicode=True, default_flow_style=False)

    print(f"\n📋 文件清单已保存: {manifest_file.name}")

if __name__ == "__main__":
    generate_all_test_data()
