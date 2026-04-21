"""
业务验收用例预期输出生成脚本
根据测试数据生成各用例的预期处理结果
"""

import os
import polars as pl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 测试数据路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "../业务验收/测试数据")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "../业务验收/预期结果")

# 字典映射
DICT_A1 = {  # 客户来源
    "线上推广": "SRC_ONLINE",
    "线下活动": "SRC_OFFLINE",
    "电话营销": "SRC_TEL",
    "自然流量": "SRC_NATURAL",
}

DICT_A2 = {  # 客户等级
    "VIP客户": "VIP",
    "普通客户": "NORMAL",
    "潜在客户": "POTENTIAL",
    "流失客户": "LOST",
}

def create_output_dir():
    """创建输出目录"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

def load_excel(path):
    """加载Excel文件"""
    full_path = os.path.join(DATA_DIR, path)
    return pl.read_excel(full_path)

def save_multi_sheet_excel(filename, sheets_data):
    """
    保存多Sheet Excel文件
    sheets_data: dict of {sheet_name: DataFrame}
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        pdf = df.to_pandas()
        for r_idx, row in enumerate(dataframe_to_rows(pdf, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    output_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(output_path)
    print(f"  生成: {filename}")
    return output_path

def generate_summary(total, success, failed, module_stats=None):
    """生成处理摘要DataFrame"""
    data = {
        "指标": ["处理时间", "总行数", "正确数据", "错误数据", "正确率"],
        "值": [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(total),
            str(success),
            str(failed),
            f"{success/total*100:.1f}%" if total > 0 else "0%"
        ]
    }
    return pl.DataFrame(data)

def add_code_columns(df):
    """添加Code列"""
    # 客户来源_Code
    df = df.with_columns(
        pl.col("客户来源").map_elements(
            lambda x: DICT_A1.get(x, "未匹配") if x else "未匹配",
            return_dtype=pl.Utf8
        ).alias("客户来源_Code")
    )
    # 客户等级_Code
    df = df.with_columns(
        pl.col("客户等级").map_elements(
            lambda x: DICT_A2.get(x, "未匹配") if x else "未匹配",
            return_dtype=pl.Utf8
        ).alias("客户等级_Code")
    )
    return df

def check_required_fields(df, field_spec):
    """检查必填字段"""
    errors = []
    valid_rows = []

    for row in df.iter_rows(named=True):
        row_errors = []
        row_num = row.get("_行号", 0)

        # 检查姓名
        if row.get("姓名") is None or str(row.get("姓名", "")).strip() == "":
            row_errors.append("姓名必填为空")

        # 检查邮箱
        if row.get("邮箱") is None or str(row.get("邮箱", "")).strip() == "":
            row_errors.append("邮箱必填为空")
        else:
            import re
            email = str(row.get("邮箱", "")).strip()
            if not re.match(r'^[\w.-]+@[\w.-]+\.\w+$', email):
                row_errors.append("邮箱格式校验失败")

        # 检查客户来源
        if row.get("客户来源") is None or str(row.get("客户来源", "")).strip() == "":
            row_errors.append("客户来源必填为空")

        if row_errors:
            errors.append({
                "_行号": row_num,
                "_错误类型": "/".join(row_errors),
                "_错误值": str(row.get("邮箱", "")),
                **row
            })
        else:
            valid_rows.append(row)

    return valid_rows, errors

def check_dup_in_list(df, check_list_df, col_name="邮箱"):
    """检查重复并返回标注"""
    check_emails = set()
    if check_list_df is not None:
        for row in check_list_df.iter_rows():
            email = str(row[check_list_df.columns.index(col_name)]).strip().lower()
            if email:
                check_emails.add(email)

    result = []
    for row in df.iter_rows(named=True):
        email = str(row.get(col_name, "")).strip().lower()
        result.append("是" if email in check_emails else "否")
    return result

# ============================================================
# BAT-01: 正常全流程处理
# ============================================================
def generate_bat01():
    """BAT-01: 正常全流程处理"""
    print("\n=== 生成 BAT-01 预期结果 ===")

    df_yixian = load_excel("一线名单_正常.xlsx")
    df_sanfang = load_excel("三方名单_正常.xlsx")
    df_hw = load_excel("HW名单_正常.xlsx")

    # 添加行号
    df_yixian = df_yixian.with_columns(pl.arange(2, len(df_yixian) + 2).alias("_行号"))

    # F4: 添加Code列
    df_yixian = add_code_columns(df_yixian)

    # 重新排列列顺序
    yixian_cols = ["_行号", "姓名", "邮箱", "手机", "客户来源", "客户来源_Code", "客户等级", "客户等级_Code"]
    df_yixian = df_yixian.select([c for c in yixian_cols if c in df_yixian.columns])

    # 生成一线名单输出
    sheets = {
        "处理摘要": generate_summary(5, 5, 0),
        "原始数据": df_yixian,
        "合规性检查结果": pl.DataFrame({"说明": ["无不合规数据"]}),
        "字典校验结果": pl.DataFrame({"说明": ["无未匹配数据"]}),
        "重复名单结果": pl.DataFrame({"说明": ["无重复数据"]}),
    }
    save_multi_sheet_excel("BAT-01_一线名单_预期结果.xlsx", sheets)

    # 生成三方名单输出（标注是否在一线）
    df_sanfang = df_sanfang.with_columns(
        pl.Series("是否已在一线名单", check_dup_in_list(df_sanfang, df_yixian))
    )
    sheets = {
        "原始数据": df_sanfang
    }
    save_multi_sheet_excel("BAT-01_三方名单_预期结果.xlsx", sheets)

    # 生成HW名单输出（双标注）
    hw_mark1 = check_dup_in_list(df_hw, df_yixian)
    hw_mark2 = check_dup_in_list(df_hw, df_sanfang)
    df_hw = df_hw.with_columns(
        pl.Series("是否已在一线名单", hw_mark1),
        pl.Series("是否已在三方名单", hw_mark2)
    )
    sheets = {
        "原始数据": df_hw
    }
    save_multi_sheet_excel("BAT-01_HW名单_预期结果.xlsx", sheets)

# ============================================================
# BAT-02: 必填字段为空
# ============================================================
def generate_bat02():
    """BAT-02: 必填字段为空"""
    print("\n=== 生成 BAT-02 预期结果 ===")

    df = load_excel("一线名单_必填空.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))

    valid_rows, errors = check_required_fields(df, None)

    valid_df = pl.DataFrame(valid_rows) if valid_rows else pl.DataFrame()
    error_df = pl.DataFrame(errors) if errors else pl.DataFrame()

    if "_行号" in valid_df.columns:
        valid_df = valid_df.select(["_行号", "姓名", "邮箱", "手机", "客户来源", "客户等级"])

    # 错误记录格式
    if error_df.columns:
        error_records = []
        for row in error_df.iter_rows(named=True):
            err_types = row.get("_错误类型", "").split("/")
            err_val = str(row.get("邮箱", ""))
            row_num = row.get("_行号", 0)

            for err_type in err_types:
                if err_type:
                    error_records.append({
                        "字段名": err_type.split("必填")[0] if "必填" in err_type else err_type.split("格式")[0],
                        "行号": row_num,
                        "问题类型": err_type,
                        "实际值": err_val
                    })

        error_result_df = pl.DataFrame(error_records)
    else:
        error_result_df = pl.DataFrame()

    sheets = {
        "处理摘要": generate_summary(5, 2, 3),
        "原始数据": valid_df if valid_df.columns else pl.DataFrame(),
        "合规性检查结果": error_result_df,
    }
    save_multi_sheet_excel("BAT-02_预期结果.xlsx", sheets)

# ============================================================
# BAT-03: 邮箱格式非法
# ============================================================
def generate_bat03():
    """BAT-03: 邮箱格式非法"""
    print("\n=== 生成 BAT-03 预期结果 ===")

    df = load_excel("一线名单_邮箱非法.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))

    valid_rows = []
    error_records = []

    for row in df.iter_rows(named=True):
        row_num = row.get("_行号", 0)
        email = str(row.get("邮箱", "")).strip()

        import re
        if not re.match(r'^[\w.-]+@[\w.-]+\.\w+$', email):
            error_records.append({
                "字段名": "邮箱",
                "行号": row_num,
                "问题类型": "邮箱格式校验失败",
                "实际值": email
            })
        else:
            valid_rows.append(row)

    valid_df = pl.DataFrame(valid_rows) if valid_rows else pl.DataFrame()
    error_df = pl.DataFrame(error_records)

    if valid_df.columns:
        valid_df = valid_df.select(["_行号", "姓名", "邮箱", "手机", "客户来源", "客户等级"])

    sheets = {
        "处理摘要": generate_summary(4, 2, 2),
        "原始数据": valid_df,
        "合规性检查结果": error_df,
    }
    save_multi_sheet_excel("BAT-03_预期结果.xlsx", sheets)

# ============================================================
# BAT-04: 字典上码正常匹配
# ============================================================
def generate_bat04():
    """BAT-04: 字典上码正常匹配"""
    print("\n=== 生成 BAT-04 预期结果 ===")

    df = load_excel("一线名单_字典正常.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))
    df = add_code_columns(df)

    # 排列列顺序
    cols = ["_行号", "姓名", "邮箱", "手机", "客户来源", "客户来源_Code", "客户等级", "客户等级_Code"]
    df = df.select([c for c in cols if c in df.columns])

    sheets = {
        "原始数据": df,
    }
    save_multi_sheet_excel("BAT-04_预期结果.xlsx", sheets)

# ============================================================
# BAT-05: 字典上码存在未匹配值
# ============================================================
def generate_bat05():
    """BAT-05: 字典上码存在未匹配值"""
    print("\n=== 生成 BAT-05 预期结果 ===")

    df = load_excel("一线名单_字典非法值.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))
    df = add_code_columns(df)

    # 找出未匹配记录
    dict_error_records = []
    for row in df.iter_rows(named=True):
        row_num = row.get("_行号", 0)
        source_code = row.get("客户来源_Code", "")
        level_code = row.get("客户等级_Code", "")

        if source_code == "未匹配":
            dict_error_records.append({
                "字段名": "客户来源",
                "行号": row_num,
                "原始值": row.get("客户来源", "")
            })
        if level_code == "未匹配":
            dict_error_records.append({
                "字段名": "客户等级",
                "行号": row_num,
                "原始值": row.get("客户等级", "")
            })

    dict_error_df = pl.DataFrame(dict_error_records)

    # 排列列
    cols = ["_行号", "姓名", "邮箱", "手机", "客户来源", "客户来源_Code", "客户等级", "客户等级_Code"]
    df = df.select([c for c in cols if c in df.columns])

    sheets = {
        "原始数据": df,
        "字典校验结果": dict_error_df,
    }
    save_multi_sheet_excel("BAT-05_预期结果.xlsx", sheets)

# ============================================================
# BAT-06: 名单内部重复检查
# ============================================================
def generate_bat06():
    """BAT-06: 名单内部重复检查"""
    print("\n=== 生成 BAT-06 预期结果 ===")

    df = load_excel("一线名单_内部重复.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))

    # 按邮箱分组找重复
    email_groups = {}
    for row in df.iter_rows(named=True):
        email = str(row.get("邮箱", "")).strip().lower()
        if email:
            if email not in email_groups:
                email_groups[email] = []
            email_groups[email].append(row)

    # 生成重复结果
    dup_records = []
    for email, rows in email_groups.items():
        if len(rows) > 1:
            for idx, row in enumerate(rows):
                dup_records.append({
                    "_行号": row.get("_行号", 0),
                    "_重复键值": email,
                    "_出现次数": len(rows),
                    "_重复标记": "原始" if idx == 0 else "重复"
                })

    dup_df = pl.DataFrame(dup_records)

    sheets = {
        "重复名单结果": dup_df,
    }
    save_multi_sheet_excel("BAT-06_预期结果.xlsx", sheets)

# ============================================================
# BAT-07: 空值不参与重复检查
# ============================================================
def generate_bat07():
    """BAT-07: 空值不参与重复检查"""
    print("\n=== 生成 BAT-07 预期结果 ===")

    df = load_excel("一线名单_邮箱空值.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))

    # 只检查非空邮箱的重复
    email_groups = {}
    for row in df.iter_rows(named=True):
        email = str(row.get("邮箱", "")).strip().lower()
        if email:  # 只处理非空邮箱
            if email not in email_groups:
                email_groups[email] = []
            email_groups[email].append(row)

    # 生成重复结果
    dup_records = []
    for email, rows in email_groups.items():
        if len(rows) > 1:
            for idx, row in enumerate(rows):
                dup_records.append({
                    "_行号": row.get("_行号", 0),
                    "_重复键值": email,
                    "_出现次数": len(rows),
                    "_重复标记": "原始" if idx == 0 else "重复"
                })

    dup_df = pl.DataFrame(dup_records)

    sheets = {
        "重复名单结果": dup_df,
    }
    save_multi_sheet_excel("BAT-07_预期结果.xlsx", sheets)

# ============================================================
# BAT-08: 跨名单去重-三方标注
# ============================================================
def generate_bat08():
    """BAT-08: 跨名单去重-三方标注"""
    print("\n=== 生成 BAT-08 预期结果 ===")

    df_yixian = load_excel("一线名单_正常.xlsx")
    df_sanfang = load_excel("三方名单_重复.xlsx")

    # 收集一线邮箱
    yixian_emails = set()
    for row in df_yixian.iter_rows():
        email = str(row[1]).strip().lower()  # 邮箱在第2列
        if email:
            yixian_emails.add(email)

    # 标注三方
    marks = []
    for row in df_sanfang.iter_rows():
        email = str(row[1]).strip().lower()  # 邮箱在第2列
        marks.append("是" if email in yixian_emails else "否")

    df_sanfang = df_sanfang.with_columns(pl.Series("是否已在一线名单", marks))

    sheets = {
        "原始数据": df_sanfang,
    }
    save_multi_sheet_excel("BAT-08_预期结果.xlsx", sheets)

# ============================================================
# BAT-09: 跨名单去重-HW双标注
# ============================================================
def generate_bat09():
    """BAT-09: 跨名单去重-HW双标注"""
    print("\n=== 生成 BAT-09 预期结果 ===")

    df_yixian = load_excel("一线名单_正常.xlsx")
    df_sanfang = load_excel("三方名单_正常.xlsx")
    df_hw = load_excel("HW名单_重复.xlsx")

    # 收集邮箱
    yixian_emails = set(str(row[1]).strip().lower() for row in df_yixian.iter_rows() if row[1])
    sanfang_emails = set(str(row[1]).strip().lower() for row in df_sanfang.iter_rows() if row[1])

    # 标注HW
    marks1, marks2 = [], []
    for row in df_hw.iter_rows():
        email = str(row[1]).strip().lower()
        marks1.append("是" if email in yixian_emails else "否")
        marks2.append("是" if email in sanfang_emails else "否")

    df_hw = df_hw.with_columns(
        pl.Series("是否已在一线名单", marks1),
        pl.Series("是否已在三方名单", marks2)
    )

    sheets = {
        "原始数据": df_hw,
    }
    save_multi_sheet_excel("BAT-09_预期结果.xlsx", sheets)

# ============================================================
# BAT-10: 去重大小写不敏感
# ============================================================
def generate_bat10():
    """BAT-10: 去重大小写不敏感"""
    print("\n=== 生成 BAT-10 预期结果 ===")

    df_yixian = load_excel("一线名单_正常.xlsx")
    df_sanfang = load_excel("三方名单_大小写.xlsx")

    # 收集一线邮箱（小写）
    yixian_emails = set(str(row[1]).strip().lower() for row in df_yixian.iter_rows() if row[1])

    # 标注三方（自动去除首尾空格并转小写比较）
    marks = []
    for row in df_sanfang.iter_rows():
        email = str(row[1]).strip().lower()  # 去除空格并转小写
        marks.append("是" if email in yixian_emails else "否")

    df_sanfang = df_sanfang.with_columns(pl.Series("是否已在一线名单", marks))

    sheets = {
        "原始数据": df_sanfang,
    }
    save_multi_sheet_excel("BAT-10_预期结果.xlsx", sheets)

# ============================================================
# BAT-11: 混合场景
# ============================================================
def generate_bat11():
    """BAT-11: 混合场景"""
    print("\n=== 生成 BAT-11 预期结果 ===")

    df = load_excel("一线名单_混合.xlsx")
    df = df.with_columns(pl.arange(2, len(df) + 2).alias("_行号"))

    # F2: 字段合规性检查
    valid_rows, f2_errors = [], []
    import re
    for row in df.iter_rows(named=True):
        row_num = row.get("_行号", 0)
        errors = []

        if not row.get("姓名", ""): errors.append("姓名必填为空")
        if not row.get("邮箱", ""): errors.append("邮箱必填为空")
        else:
            email = str(row.get("邮箱", "")).strip()
            if not re.match(r'^[\w.-]+@[\w.-]+\.\w+$', email):
                errors.append("邮箱格式校验失败")
        if not row.get("客户来源", ""): errors.append("客户来源必填为空")
        if row.get("手机"):
            phone = str(row.get("手机", "")).strip()
            if phone and not re.match(r'^1[3-9]\d{9}$', phone):
                errors.append("手机格式校验失败")

        if errors:
            for err in errors:
                field = err.split("必填")[0].split("格式")[0]
                f2_errors.append({
                    "字段名": field,
                    "行号": row_num,
                    "问题类型": err,
                    "实际值": str(row.get(field if field != "邮箱" else "邮箱", ""))
                })
        else:
            valid_rows.append(row)

    # F6: 内部重复检查（只对有效行）
    valid_df = pl.DataFrame(valid_rows) if valid_rows else pl.DataFrame()
    dup_records = []
    if "_行号" in valid_df.columns:
        email_groups = {}
        for row in valid_df.iter_rows(named=True):
            email = str(row.get("邮箱", "")).strip().lower()
            if email:
                if email not in email_groups:
                    email_groups[email] = []
                email_groups[email].append(row)

        for email, rows in email_groups.items():
            if len(rows) > 1:
                for idx, row in enumerate(rows):
                    dup_records.append({
                        "_行号": row.get("_行号", 0),
                        "_重复键值": email,
                        "_出现次数": len(rows),
                        "_重复标记": "原始" if idx == 0 else "重复"
                    })

    # F4: 添加Code列（只对有效行）
    if valid_rows:
        valid_df = add_code_columns(valid_df)

    # F5: 字典校验（只对有效行）
    f5_errors = []
    if valid_rows:
        for row in valid_df.iter_rows(named=True):
            if row.get("客户来源_Code") == "未匹配":
                f5_errors.append({
                    "字段名": "客户来源",
                    "行号": row.get("_行号", 0),
                    "原始值": row.get("客户来源", "")
                })
            if row.get("客户等级_Code") == "未匹配":
                f5_errors.append({
                    "字段名": "客户等级",
                    "行号": row.get("_行号", 0),
                    "原始值": row.get("客户等级", "")
                })

    # 排列列
    cols = ["_行号", "姓名", "邮箱", "手机", "客户来源", "客户来源_Code", "客户等级", "客户等级_Code"]
    valid_df = valid_df.select([c for c in cols if c in valid_df.columns])

    sheets = {
        "处理摘要": generate_summary(8, 5, 3),
        "原始数据": valid_df,
        "合规性检查结果": pl.DataFrame(f2_errors),
        "字典校验结果": pl.DataFrame(f5_errors),
        "重复名单结果": pl.DataFrame(dup_records),
    }
    save_multi_sheet_excel("BAT-11_预期结果.xlsx", sheets)

# ============================================================
# 主函数
# ============================================================
def main():
    print("=" * 60)
    print("业务验收用例预期结果生成")
    print("=" * 60)

    create_output_dir()

    generate_bat01()  # 正常全流程
    generate_bat02()  # 必填为空
    generate_bat03()  # 邮箱非法
    generate_bat04()  # 字典上码正常
    generate_bat05()  # 字典上码未匹配
    generate_bat06()  # 内部重复
    generate_bat07()  # 空值不参与重复
    generate_bat08()  # 三方标注
    generate_bat09()  # HW双标注
    generate_bat10()  # 大小写不敏感
    generate_bat11()  # 混合场景

    print("\n" + "=" * 60)
    print("生成完成！输出目录:", OUTPUT_DIR)
    print("=" * 60)

if __name__ == "__main__":
    main()
