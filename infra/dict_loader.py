"""
基础设施层 - 字典加载器（F4-02 / F5 核心基础设施）

负责加载并解析 data_dict.xlsx（租户字典导入模版），
构建内存中的字典映射结构供 F4/F5 模块使用。

字典文件格式（已确认 DEV-06）：
- 单 Sheet 结构
- 横排格式：Code 列 + 相邻右列 Details 列构成一组
  （第0行用于标识字典类型）
- 支持 MD5 版本感知，可检测字典变更

内存映射结构：
{
  "字典名1": {
    "代码值1": "显示值1",
    "代码值2": "显示值2",
    ...
  },
  ...
}
"""

import hashlib
import os
from typing import Any, Dict, List, Optional

import openpyxl

from infra.exceptions import DataQualityError, ValidationError
from infra.log_manager import get_logger

logger = get_logger(__name__)


class DictLoader:
    """
    字典加载器

    Attributes
    ----------
    file_path : str
        字典文件路径
    md5_hash : str
        当前字典文件的 MD5 哈希值
    mappings : Dict[str, Dict[str, str]]
        内存字典映射 {字典名: {代码值: 显示值}}
    """

    def __init__(self, file_path: str):
        if not os.path.exists(file_path):
            raise DataQualityError(
                f"字典文件不存在: {file_path}",
                field="data_dict_path",
            )
        self.file_path = file_path
        self.md5_hash: Optional[str] = None
        self.mappings: Dict[str, Dict[str, str]] = {}
        self._dict_names: List[str] = []
        self._load()

    def _load(self):
        """内部加载逻辑"""
        self.md5_hash = self._compute_md5()
        workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet = workbook.active

        # 读取表头行（3行）
        header_rows = []
        for row_idx in range(1, min(4, sheet.max_row + 1)):
            header_rows.append([cell.value for cell in sheet[row_idx]])

        dict_groups = self._parse_header(header_rows)

        # 逐组解析数据（从第4行开始，跳过3行表头）
        for dict_name, code_col, details_col in dict_groups:
            code_to_label: Dict[str, str] = {}
            for row in sheet.iter_rows(min_row=4, values_only=True):  # 数据从第4行开始
                code_val = row[code_col] if code_col < len(row) else None
                label_val = row[details_col] if details_col < len(row) else None
                if code_val is None and label_val is None:
                    continue
                code_str = str(code_val).strip() if code_val is not None else ""
                label_str = str(label_val).strip() if label_val is not None else ""
                if code_str:
                    code_to_label[code_str] = label_str

            self.mappings[dict_name] = code_to_label
            self._dict_names.append(dict_name)

        workbook.close()
        logger.info(
            f"字典加载完成: {self.file_path}, "
            f"共 {len(self.mappings)} 个字典, MD5={self.md5_hash}"
        )

    def _parse_header(self, header_rows: List[List]) -> List[tuple]:
        """
        解析表头行，识别字典组。

        支持两种字典文件格式：
        - 模版格式（第1行有"字典类型标识"标签）：
          行1: "字典类型标识", "示例：A1", ...
          行2: "字典类型名称", "示例：客户等级", ...
          行3: Code, Details, ...
        - 业务格式（第1行直接是dict_id）：
          行1: "A1", None, "A2", None
          行2: "客户等级", None, "客户来源", None
          行3: Code, Details, ...

        Returns
        -------
        List[(dict_name, code_col_idx, details_col_idx)]
        """
        groups = []
        col_count = len(header_rows[0])

        row3_col_header = header_rows[2]  # 第3行：Code/Details列头

        # 判断格式：第1行是否包含"字典类型标识"
        row1_has_label = any("字典类型标识" in str(v) for v in header_rows[0] if v)

        if row1_has_label:
            # 模版格式：从第2行获取dict_id
            dict_id_row = header_rows[1]
        else:
            # 业务格式：从第1行获取dict_id
            dict_id_row = header_rows[0]

        # 扫描第3行的Code列（Code列在左，Details列在右）
        for col in range(col_count):
            code_header = row3_col_header[col] if col < len(row3_col_header) else None
            if code_header is None:
                continue

            # 检查是否为Code列
            if not ("code" in str(code_header).lower() or "编码" in str(code_header)):
                continue

            details_col = col + 1
            if details_col >= col_count:
                continue

            details_header = row3_col_header[details_col] if details_col < len(row3_col_header) else None
            if details_header is None:
                continue

            # 验证Details列
            if not ("details" in str(details_header).lower() or "说明" in str(details_header) or "标签" in str(details_header)):
                continue

            # 从dict_id_row获取dict_id
            code_cell_val = dict_id_row[col] if col < len(dict_id_row) else None
            dict_id = None
            if code_cell_val:
                dict_id = self._extract_dict_id(str(code_cell_val).strip())

            # 如果提取失败，使用列索引作为fallback
            if not dict_id:
                dict_id = f"Dict_{col}"

            groups.append((dict_id, col, details_col))
            logger.info(f"[dict_loader] 解析到字典: {dict_id}, Code列={col}, Details列={details_col}")

        if not groups:
            raise ValidationError(
                f"字典文件格式无法解析。请确认文件格式符合规范。",
                code="DICT_HEADER",
            )

        return groups

    def _extract_dict_id(self, cell_str: str) -> str:
        """
        从单元格文本中提取纯净的dict_id。

        支持格式：
        - "示例：A1" -> "A1"
        - "A1" -> "A1"
        """
        # 移除常见前缀
        prefixes_to_remove = ["示例：", "示例:", "Dict:", "Dict："]
        result = cell_str
        for prefix in prefixes_to_remove:
            if result.startswith(prefix):
                result = result[len(prefix):]
                break
        return result.strip()

    def _parse_header_compat(self, header_row: List) -> List[tuple]:
        """
        [兼容性] 旧版表头解析逻辑：从列名含 "Code"/"编码" 识别字典组。

        Returns
        -------
        List[(dict_name, code_col_idx, details_col_idx)]
        """
        groups = []
        col_count = len(header_row)
        col = 0

        while col < col_count - 1:
            header_val = header_row[col]
            if header_val is None:
                col += 1
                continue

            header_str = str(header_val).strip()
            # 判断是否为 Code 列
            if "code" in header_str.lower() or "编码" in header_str:
                dict_name = self._derive_dict_name(header_str, col, header_row)
                # 右侧相邻列默认为 Details
                details_col = col + 1
                if details_col < col_count:
                    groups.append((dict_name, col, details_col))
                    col += 2
                    continue
            col += 1

        return groups

    def _derive_dict_name(self, header_str: str, col_idx: int, header_row: List) -> str:
        """
        从表头文本推断字典名称（兼容性方法）。

        规则：取 Code 前的文本作为字典名；
        若无前序文本，使用列所在组的上方单元格或列索引。
        """
        # 移除 "Code" / "code" 等后缀
        name = header_str
        for suffix in ["Code", "CODE", "代码", "编码"]:
            if name.endswith(suffix):
                name = name[: -len(suffix)].strip()
                break

        if name:
            return name

        # 检查右侧相邻列的表头（通常标注为 Details / 说明）
        if col_idx + 1 < len(header_row):
            next_header = header_row[col_idx + 1]
            if next_header:
                next_str = str(next_header).strip()
                for suffix in ["Details", "DETAILS", "说明", "Label", "标签"]:
                    if next_str.endswith(suffix):
                        return next_str[: -len(suffix)].strip()

        # 回退：使用列字母
        return f"Dict_Col{col_idx}"

    def _compute_md5(self) -> str:
        """计算字典文件的 MD5 哈希值"""
        md5 = hashlib.md5()
        with open(self.file_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                md5.update(chunk)
        return md5.hexdigest()

    def is_code_valid(self, dict_name: str, code_value: str) -> bool:
        """判断代码值是否在指定字典中"""
        mapping = self.mappings.get(dict_name, {})
        return code_value in mapping

    def lookup(self, dict_name: str, code_value: str) -> Optional[str]:
        """
        根据字典名和代码值查询显示值。

        Returns
        -------
        str or None（代码值不存在时返回 None）
        """
        mapping = self.mappings.get(dict_name, {})
        return mapping.get(code_value)

    def get_all_codes(self, dict_name: str) -> List[str]:
        """获取指定字典的所有代码值列表"""
        mapping = self.mappings.get(dict_name, {})
        return list(mapping.keys())

    def get_all_dict_names(self) -> List[str]:
        """获取所有已加载的字典名称列表"""
        return list(self._dict_names)

    def filter_valid_rows(
        self,
        df,
        dict_name: str,
        code_column: str,
        error_column: str = None,
    ):
        """
        从 DataFrame 中筛选出字典值合法的行，
        并将非法行的错误信息写入指定列。

        Parameters
        ----------
        df : polars.DataFrame
            输入数据
        dict_name : str
            字典名称
        code_column : str
            代码值所在列名
        error_column : str, optional
            错误信息写入列名

        Returns
        -------
        polars.DataFrame
        """
        import polars as pl

        mapping = self.mappings.get(dict_name, {})

        def check_code(code) -> tuple:
            code_str = str(code) if code is not None else ""
            if not code_str:
                return (True, None)
            is_valid = code_str in mapping
            return (is_valid, None if is_valid else f"代码值 '{code_str}' 不在字典 {dict_name} 中")

        if error_column:
            return df.with_columns(
                pl.col(code_column).map_elements(
                    lambda x: check_code(x)[1],
                    return_dtype=pl.String,
                ).alias(error_column)
            )
        return df.filter(
            pl.col(code_column).map_elements(
                lambda x: check_code(x)[0],
                return_dtype=pl.Boolean,
            )
        )
