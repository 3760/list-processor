"""
基础设施层 - Excel 写入工具

封装 openpyxl，提供便捷的多 Sheet Excel 写入能力，
支持追加 Sheet、样式设置、行列数据写入等功能。
"""

import os
from typing import Any, Dict, List, Optional, Union

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from infra.exceptions import DataQualityError
from infra.log_manager import get_logger

logger = get_logger(__name__)


class ExcelWriter:
    """
    Excel 写入工具类。

    Usage:
        writer = ExcelWriter("output.xlsx")
        writer.write_sheet("Sheet1", headers, data_rows)
        writer.write_sheet("异常数据", headers2, error_rows)
        writer.save()
    """

    def __init__(self, file_path: str, overwrite: bool = True):
        self.file_path = file_path
        if overwrite and os.path.exists(file_path):
            os.remove(file_path)

        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # 删除默认 Sheet
        self._sheet_count = 0

    def write_sheet(
        self,
        sheet_name: str,
        headers: List[str],
        rows: List[List[Any]],
        column_widths: Optional[List[int]] = None,
        freeze_panes: str = "A2",
    ):
        """
        向 Excel 中写入一个 Sheet。

        Parameters
        ----------
        sheet_name : str
            Sheet 名称
        headers : List[str]
            表头行（作为第1行写入）
        rows : List[List]
            数据行列表
        column_widths : List[int], optional
            各列宽度（字符数）
        freeze_panes : str, optional
            冻结窗格位置，默认为 "A2"（冻结首行）
        """
        # Sheet 名称去重
        safe_name = self._safe_sheet_name(sheet_name)
        ws = self.workbook.create_sheet(safe_name)
        self._sheet_count += 1

        # 写入表头（带样式）
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        # 写入数据行
        for row in rows:
            ws.append(list(row))

        # 设置列宽
        if column_widths:
            for col_idx, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width
        else:
            # 自动列宽
            for col_idx, cell in enumerate(ws[1], start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    len(str(cell.value or "")) + 4, 10
                )

        # 冻结首行
        if freeze_panes:
            ws.freeze_panes = freeze_panes

        logger.debug(f"写入 Sheet '{safe_name}': {len(headers)} 列, {len(rows)} 行")

    def _style_header_row(self, ws, num_cols: int):
        """为表头行设置样式"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 20

    def _safe_sheet_name(self, name: str) -> str:
        """生成合法的 Sheet 名称（Excel 限制31字符以内）"""
        safe = name[:31].strip()
        if not safe:
            safe = f"Sheet{self._sheet_count + 1}"
        # Sheet 名称不能包含 : / \ * ? [ ]
        for char in [":", "/", "\\", "*", "?", "[", "]"]:
            safe = safe.replace(char, "_")
        return safe

    def append_sheet(
        self,
        sheet_name: str,
        headers: List[str],
        rows: List[List[Any]],
    ):
        """
        追加数据到已存在的同名 Sheet（不存在则创建）。
        用于将多个来源的数据合并到同一 Sheet。
        """
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            # 追加行（跳过表头）
            for row in rows:
                ws.append(list(row))
            logger.debug(f"追加到 Sheet '{sheet_name}': {len(rows)} 行")
        else:
            self.write_sheet(sheet_name, headers, rows)

    def save(self):
        """保存文件到磁盘"""
        self.workbook.save(self.file_path)
        logger.info(f"Excel 已保存: {self.file_path}")

    def close(self):
        """关闭工作簿"""
        self.workbook.close()
