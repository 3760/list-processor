"""
Tests for F1: File Loader Module

F1 功能覆盖：
- F1-01 ~ F1-04: 文件加载、Sheet选择、_来源列添加
- F1-05: 多Sheet选择对话框（DEV-05）
- F1-06: 去重字段识别（三级策略，NQ-02）
- F1-07: 字段规范导入对话框（T-23，不在F1模块内）
- F1-08: attr_code 唯一性预检验
- F1-09: 加载异常处理
- F1-10: 空文件处理

PROJ-04 三级去重字段识别策略：
- Level 1: config 读取（由调用方保证传入）
- Level 2: 关键字模糊匹配
- Level 3: 弹窗手动指定（UI层处理，本测试覆盖候选列提供）
"""

from pathlib import Path

import openpyxl
import polars as pl
import pytest

from core.context import ProcessContext
from infra.exceptions import DataQualityError
from modules.f1_loader import (
    detect_dedup_field,
    get_all_columns,
    load_files,
)


# ─────────────────────────────────────────────
# 辅助函数：生成测试 Excel 文件
# ─────────────────────────────────────────────

def _make_excel(path: Path, data: list[dict], sheet_name: str = "Sheet1"):
    """生成测试用 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row.get(h) for h in headers])
    wb.save(path)
    wb.close()


def _make_multi_sheet_excel(path: Path, sheets: list[tuple[str, list[dict]]]):
    """生成多Sheet Excel 文件"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, data in sheets:
        ws = wb.create_sheet(title=sheet_name)
        if data:
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h) for h in headers])
    wb.save(path)
    wb.close()


# ─────────────────────────────────────────────
# F1-01 ~ F1-04: 文件加载核心逻辑
# ─────────────────────────────────────────────

class TestFileLoading:
    """F1-01~F1-04: 文件加载、_来源列添加"""

    def test_load_single_sheet_adds_source_column(self, tmp_path):
        """F1-01: 单Sheet文件加载，自动添加 `_来源` 列"""
        path = tmp_path / "名单.xlsx"
        _make_excel(path, [
            {"姓名": "张三", "邮箱": "zhang@example.com"},
            {"姓名": "李四", "邮箱": "li@example.com"},
        ])
        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": str(path)})

        df = ctx.get_dataframe("yixian")
        assert df is not None
        assert len(df) == 2
        assert "_来源" in df.columns
        assert df["_来源"].to_list() == ["一线", "一线"]

    def test_load_multiple_list_types(self, tmp_path):
        """F1-02: 多类型名单加载（同时加载一线/三方/HW）"""
        fp_yixian = tmp_path / "一线.xlsx"
        fp_sanfang = tmp_path / "三方.xlsx"
        fp_hw = tmp_path / "HW.xlsx"

        _make_excel(fp_yixian, [{"邮箱": "a@x.com"}])
        _make_excel(fp_sanfang, [{"邮箱": "b@x.com"}])
        _make_excel(fp_hw, [{"邮箱": "c@x.com"}])

        ctx = ProcessContext()
        ctx = load_files(ctx, {
            "一线": str(fp_yixian),
            "三方": str(fp_sanfang),
            "HW": str(fp_hw),
        })

        assert ctx.get_dataframe("yixian") is not None
        assert ctx.get_dataframe("sanfang") is not None
        # 注意：内部 key 大小写敏感，"HW" 与 "hw" 不同
        assert ctx.get_dataframe("HW") is not None

    def test_skip_missing_file(self, tmp_path):
        """F1-03: 未指定文件路径时跳过，不抛出异常"""
        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": "", "三方": None})
        assert ctx.get_dataframe("yixian") is None
        assert ctx.get_dataframe("sanfang") is None

    def test_empty_file_raises_error(self, tmp_path):
        """F1-09: 空文件（无数据行）抛出 DataQualityError"""
        path = tmp_path / "empty.xlsx"
        _make_excel(path, [])
        ctx = ProcessContext()
        with pytest.raises(DataQualityError, match="empty"):
            load_files(ctx, {"一线": str(path)})

    def test_nonexistent_file_raises_error(self, tmp_path):
        """F1-09: 文件不存在抛出 DataQualityError（实际抛 ValidationError 继承类）"""
        ctx = ProcessContext()
        with pytest.raises(DataQualityError, match="不存在"):
            load_files(ctx, {"一线": "/fake/nonexistent/file.xlsx"})

    def test_dedup_field_recorded_in_context(self, tmp_path):
        """F1-06: 去重字段记录到 ProcessContext"""
        path = tmp_path / "名单.xlsx"
        _make_excel(path, [{"邮箱": "a@x.com"}])
        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": str(path)}, dedup_field="邮箱")
        assert ctx.dedup_field == "邮箱"


# ─────────────────────────────────────────────
# DEV-05: 多Sheet选择策略
# ─────────────────────────────────────────────

class TestMultiSheetSelection:
    """F1-05: 多Sheet选择对话框（DEV-05）"""

    def test_single_sheet_no_callback(self, tmp_path):
        """Sheet=1：直接读取，无提示"""
        path = tmp_path / "single.xlsx"
        _make_excel(path, [{"邮箱": "a@x.com"}])

        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": str(path)}, sheet_selection_callback=None)

        df = ctx.get_dataframe("yixian")
        assert df is not None
        assert len(df) == 1

    def test_multi_sheet_with_callback_user_selects(self, tmp_path):
        """Sheet>1：用户回调选择 Sheet2"""
        path = tmp_path / "multi.xlsx"
        _make_multi_sheet_excel(path, [
            ("Sheet1", [{"邮箱": "from_sheet1@test.com"}]),
            ("Sheet2", [{"邮箱": "from_sheet2@test.com"}]),
            ("Sheet3", [{"邮箱": "from_sheet3@test.com"}]),
        ])

        def cb(names):
            assert "Sheet1" in names
            assert "Sheet2" in names
            assert "Sheet3" in names
            return "Sheet2"

        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": str(path)}, sheet_selection_callback=cb)

        df = ctx.get_dataframe("yixian")
        assert df is not None
        assert len(df) == 1
        assert df["邮箱"][0] == "from_sheet2@test.com"

    def test_multi_sheet_callback_returns_none_uses_first(self, tmp_path):
        """Sheet>1：回调返回 None（超时），默认选第一个Sheet"""
        path = tmp_path / "multi.xlsx"
        _make_multi_sheet_excel(path, [
            ("Sheet1", [{"邮箱": "first@test.com"}]),
            ("Sheet2", [{"邮箱": "second@test.com"}]),
        ])

        def cb(names):
            return None  # 超时未选择

        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": str(path)}, sheet_selection_callback=cb)

        df = ctx.get_dataframe("yixian")
        assert df is not None
        assert df["邮箱"][0] == "first@test.com"


# ─────────────────────────────────────────────
# NQ-02 / PROJ-04: 三级去重字段识别策略
# ─────────────────────────────────────────────

class TestDedupFieldDetection:
    """F1-06 / PROJ-04: 三级去重字段识别策略"""

    @pytest.mark.parametrize("col_name,expected", [
        ("email", "email"),
        ("E-Mail", "E-Mail"),
        ("邮箱", "邮箱"),
        ("mail_address", "mail_address"),
        ("e_mail", "e_mail"),
        ("Phone", None),
        ("姓名", None),
    ])
    def test_detect_dedup_field_keyword_match(self, col_name, expected):
        """Level 2: 关键字匹配（email/邮箱/mail/e-mail，大小写不敏感）"""
        df = pl.DataFrame({col_name: ["a@b.com", "c@d.com"], "姓名": ["张三", "李四"]})
        result = detect_dedup_field(df)
        assert result == expected

    def test_detect_dedup_field_prefers_email_over_others(self):
        """Level 2: 多个列匹配时，取第一个匹配到（遍历顺序）"""
        df = pl.DataFrame({
            "姓名": ["张三", "李四"],
            "邮箱": ["a@b.com", "c@d.com"],
            "mail": ["e@f.com", "g@h.com"],
        })
        result = detect_dedup_field(df)
        assert result == "邮箱"

    def test_get_all_columns(self):
        """Level 3 支持：提供所有列名供用户下拉选择"""
        df = pl.DataFrame({"姓名": [], "邮箱": [], "手机": []})
        cols = get_all_columns(df)
        assert set(cols) == {"姓名", "邮箱", "手机"}


# ─────────────────────────────────────────────
# F1-10: 边界情况
# ─────────────────────────────────────────────

class TestEdgeCases:
    """F1-10: 边界情况处理"""

    def test_special_characters_in_column_names(self, tmp_path):
        """列名含特殊字符（空格、中文、括号）"""
        path = tmp_path / "special.xlsx"
        _make_excel(path, [
            {"客户邮箱（首选）": "a@b.com", "手机 号码": "13800000000"}
        ])
        ctx = ProcessContext()
        ctx = load_files(ctx, {"一线": str(path)})
        df = ctx.get_dataframe("yixian")
        assert "_来源" in df.columns
        assert "客户邮箱（首选）" in df.columns
