"""
单测 - 字段规范导入对话框解析逻辑（ui/widgets/spec_parser.py）

覆盖场景：
1. 正常解析：属性编码、属性名称、必填、数据类型、数据子类型、数据字典
2. 缺少必需列：attr_code 或 attr_name 缺失时抛出 ValidationError
3. 空文件：Excel 无有效数据行时抛出 ValidationError
4. 必填判断：多种必填标识的解析（是/否/允许空/n/no/0）
5. 文件不存在：抛出 DataQualityError
6. YAML写入：成功写入正确格式
"""

import os
import tempfile
import pytest

import openpyxl

from infra.exceptions import ValidationError, DataQualityError
from ui.widgets.spec_parser import parse_spec_excel, write_spec_yaml


# ============================================================
# Fixtures
# ============================================================

@pytest.fixture
def temp_output():
    """临时输出YAML文件路径"""
    fd, path = tempfile.mkstemp(suffix=".yaml")
    os.close(fd)
    os.remove(path)  # 清理，确保文件不存在
    yield path
    if os.path.exists(path):
        os.remove(path)


@pytest.fixture
def valid_excel():
    """有效的属性导入模版Excel（带完整列）"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "属性导入模版"

    # 表头行（符合 CEM 属性导入模版格式）
    ws.append([
        "属性编码", "属性名称", "属性类型", "数据类型",
        "数据子类型", "长度上限", "小数位数", "数据字典", "属性值必填", "验证规则"
    ])

    # 数据行
    ws.append([
        "phone", "手机号", "必填", "字符串", "", "20", "", "", "是", ""
    ])
    ws.append([
        "name", "姓名", "必填", "字符串", "", "50", "", "", "是", ""
    ])
    ws.append([
        "province", "省份", "可选", "字符串", "枚举", "", "", "A1", "否", ""
    ])
    ws.append([
        "city", "城市", "可选", "字符串", "", "30", "", "", "否", ""
    ])

    wb.save(path)
    yield path
    os.remove(path)


@pytest.fixture
def minimal_excel():
    """最小化有效Excel（只有必需列）"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["属性编码", "属性名称"])
    ws.append(["phone", "手机号"])
    ws.append(["email", "邮箱"])

    wb.save(path)
    yield path
    os.remove(path)


@pytest.fixture
def empty_excel():
    """空数据Excel（只有表头，无数据行）"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["属性编码", "属性名称"])
    # 无数据行

    wb.save(path)
    yield path
    os.remove(path)


@pytest.fixture
def no_attr_code_excel():
    """缺少属性编码列的Excel"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["属性名称", "数据类型", "属性值必填"])
    ws.append(["手机号", "字符串", "是"])

    wb.save(path)
    yield path
    os.remove(path)


@pytest.fixture
def no_attr_name_excel():
    """缺少属性名称列的Excel"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["属性编码", "数据类型", "属性值必填"])
    ws.append(["phone", "字符串", "是"])

    wb.save(path)
    yield path
    os.remove(path)


# ============================================================
# 测试用例
# ============================================================

class TestSpecImportDialogParsing:
    """字段规范解析逻辑测试"""

    def test_parse_valid_excel(self, valid_excel):
        """✅ 正常解析：完整列的Excel"""
        specs = parse_spec_excel(valid_excel)

        assert len(specs) == 4
        assert specs[0]["attr_code"] == "phone"
        assert specs[0]["attr_name"] == "手机号"
        assert specs[0]["required"] is True
        assert specs[0]["data_type"] == "字符串"

        assert specs[2]["attr_code"] == "province"
        assert specs[2]["data_subtype"] == "枚举"
        assert specs[2]["dict_id"] == "A1"
        assert specs[2]["required"] is False

    def test_parse_minimal_excel(self, minimal_excel):
        """✅ 最小列解析：只有必需列"""
        specs = parse_spec_excel(minimal_excel)

        assert len(specs) == 2
        assert specs[0]["attr_code"] == "phone"
        assert specs[0]["attr_name"] == "手机号"
        assert specs[0]["required"] is False  # 无必填列时默认可选
        assert specs[0]["data_type"] == "string"  # 默认字符串

    def test_parse_empty_excel(self, empty_excel):
        """❌ 空数据：只有表头无数据行"""
        with pytest.raises(ValidationError) as exc_info:
            parse_spec_excel(empty_excel)

        assert exc_info.value.code == "SPEC_IMPORT_EMPTY"

    def test_parse_missing_attr_code_column(self, no_attr_code_excel):
        """❌ 缺少属性编码列"""
        with pytest.raises(ValidationError) as exc_info:
            parse_spec_excel(no_attr_code_excel)

        assert exc_info.value.code == "COL_MISSING"
        assert "attr_code" in str(exc_info.value)

    def test_parse_missing_attr_name_column(self, no_attr_name_excel):
        """❌ 缺少属性名称列"""
        with pytest.raises(ValidationError) as exc_info:
            parse_spec_excel(no_attr_name_excel)

        assert exc_info.value.code == "COL_MISSING"
        assert "attr_name" in str(exc_info.value)

    def test_parse_nonexistent_file(self):
        """❌ 文件不存在"""
        with pytest.raises(DataQualityError):
            parse_spec_excel("/nonexistent/path/to/file.xlsx")


class TestRequiredFieldParsing:
    """必填标识解析测试"""

    @pytest.mark.parametrize("raw_value,expected", [
        ("是", True),
        ("1", True),
        ("true", True),
        ("否", False),
        ("允许空", False),
        ("n", False),
        ("no", False),
        ("0", False),
        ("", False),  # 空值默认 False（可选）
        (None, False),  # None 默认 False（可选）
    ])
    def test_required_parsing(self, raw_value, expected):
        """✅ 各种必填标识的解析"""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["属性编码", "属性名称", "属性值必填"])
        ws.append(["phone", "手机号", raw_value])
        wb.save(path)

        try:
            specs = parse_spec_excel(path)
            assert specs[0]["required"] == expected, f"输入值 {raw_value!r} 应解析为 {expected}"
        finally:
            os.remove(path)


class TestDataTypeParsing:
    """数据类型解析测试"""

    @pytest.mark.parametrize("raw_value,expected", [
        ("字符串", "字符串"),
        ("整数", "整数"),
        ("小数", "小数"),
        ("日期", "日期"),
        ("email", "email"),
        ("string", "string"),  # 英文输入
        ("", "string"),  # 空值默认 string
        (None, "string"),  # None 默认 string
    ])
    def test_data_type_parsing(self, raw_value, expected):
        """✅ 各种数据类型的解析"""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["属性编码", "属性名称", "数据类型"])
        ws.append(["phone", "手机号", raw_value])
        wb.save(path)

        try:
            specs = parse_spec_excel(path)
            assert specs[0]["data_type"] == expected
        finally:
            os.remove(path)


class TestWriteYaml:
    """YAML写入测试"""

    def test_write_yaml_success(self, valid_excel, temp_output):
        """✅ 正常写入YAML"""
        specs = parse_spec_excel(valid_excel)
        write_spec_yaml(specs, valid_excel, temp_output)

        assert os.path.exists(temp_output)

        import yaml
        with open(temp_output, "r", encoding="utf-8") as f:
            spec = yaml.safe_load(f)

        assert spec["version"] == "1.0"
        assert spec["source"] == valid_excel
        assert "fields" in spec
        assert len(spec["fields"]) == 4
        assert "手机号" in spec["fields"]
        assert spec["fields"]["手机号"]["attr_code"] == "phone"
        assert spec["fields"]["手机号"]["required"] is True
        # [20260420] 验证 sub_type 键名（对齐 field_spec.yaml 规范）
        assert "手机号" in spec["fields"]
        assert "type" in spec["fields"]["手机号"]  # 验证 type 而非 data_type
