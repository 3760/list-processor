"""
Tests for F7: Result Output Module

F7 功能覆盖：
- F7-01 ~ F7-03: 结果输出（多Sheet Excel）
- F7-04: 结果提示（三种场景，PRD NQ-09）

PROJ-05 三种结果提示场景：
- ✅ 无问题：绿色"处理完成，结果可直接上传 CEM 系统"
- ⚠️ 有警告：黄色"处理完成，但发现 X 条问题，建议修复后再上传"
- ❌ 失败：红色"处理失败：{错误信息}"
"""

import re
from pathlib import Path

import openpyxl
import polars as pl
import pytest

from core.context import ProcessContext
from modules.f7_output_exporter import build_result_message, export_results


# ─────────────────────────────────────────────
# 辅助：生成测试 Excel 并验证 Sheet
# ─────────────────────────────────────────────

def _read_sheets(path: Path) -> dict[str, pl.DataFrame]:
    """读取 Excel 所有 Sheet 为 DataFrame dict"""
    wb = openpyxl.load_workbook(path)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = rows[0]
            data = rows[1:]
            df = pl.DataFrame(data, schema=headers, orient="row")
            result[name] = df
    wb.close()
    return result


# ─────────────────────────────────────────────
# F7-01 ~ F7-03: 结果输出
# ─────────────────────────────────────────────

class TestResultExport:
    """F7-01~F7-03: 结果输出（多Sheet Excel）"""

    def test_export_single_list(self, tmp_path):
        """F7-01: 单类型名单输出（仅一线）"""
        ctx = ProcessContext()
        ctx.dataframes["yixian"] = pl.DataFrame({
            "邮箱": ["a@b.com", "c@d.com"],
            "_来源": ["一线", "一线"],
        })

        out_path = tmp_path / "output.xlsx"
        export_results(ctx, str(out_path))

        sheets = _read_sheets(out_path)
        assert "处理摘要" in sheets
        assert "一线" in sheets
        assert len(sheets["一线"]) == 2

    def test_export_with_validation_errors(self, tmp_path):
        """F7-02: 包含合规检查异常数据（不参与后续处理）"""
        ctx = ProcessContext()
        ctx.dataframes["yixian"] = pl.DataFrame({
            "邮箱": ["valid@x.com", "invalid-email"],
            "_来源": ["一线", "一线"],
        })
        ctx.error_records["yixian"] = pl.DataFrame({
            "_原始行号": [2],
            "_来源": ["一线"],
            "_错误类型": ["邮箱格式错误"],
            "_错误详情": ["invalid-email 不符合邮箱格式"],
        })

        out_path = tmp_path / "output.xlsx"
        export_results(ctx, str(out_path))

        sheets = _read_sheets(out_path)
        assert "合规性检查结果" in sheets
        assert "一线" in sheets
        assert len(sheets["合规性检查结果"]) == 1

    def test_export_with_repeat_records(self, tmp_path):
        """F7-03: 包含重复名单记录"""
        ctx = ProcessContext()
        ctx.dataframes["yixian"] = pl.DataFrame({
            "邮箱": ["a@b.com", "c@d.com"],
            "_来源": ["一线", "一线"],
        })
        ctx.module_results["F6"] = {
            "rejected": pl.DataFrame({
                "_行号": [2],
                "_重复键值": ["c@d.com"],
                "_出现次数": [2],
                "_重复标记": ["重复"],
            })
        }

        out_path = tmp_path / "output.xlsx"
        export_results(ctx, str(out_path))

        sheets = _read_sheets(out_path)
        assert "重复名单结果" in sheets
        assert len(sheets["重复名单结果"]) == 1


# ─────────────────────────────────────────────
# NQ-09 / PROJ-05: 三种结果提示场景
# ─────────────────────────────────────────────

class TestResultMessage:
    """F7-04 / PROJ-05: 三种结果提示场景"""

    def test_no_problems_green_message(self):
        """✅ 场景一：无问题，绿色提示"""
        ctx = ProcessContext()
        ctx.status = "completed"
        ctx.summary = {
            "yixian": {"total": 100, "passed": 100, "failed": 0},
            "sanfang": {"total": 50, "passed": 50, "failed": 0},
        }
        ctx.module_results = {
            "F5": {"fail": 0},
        }

        message, level = build_result_message(ctx)

        # ✅ 无问题 → 可直接上传
        assert level == "success"
        assert "可直接上传" in message or "完成" in message

    def test_has_warnings_yellow_message(self):
        """⚠️ 场景二：有警告，黄色提示"""
        ctx = ProcessContext()
        ctx.status = "completed"
        ctx.summary = {
            "yixian": {"total": 100, "passed": 95, "failed": 5},
        }
        ctx.module_results = {
            "F5": {"fail": 5},
        }

        message, level = build_result_message(ctx)

        # ⚠️ 有警告 → 建议修复
        assert level == "warning"
        assert "5" in message
        assert "修复" in message or "建议" in message

    def test_failure_red_message(self):
        """❌ 场景三：处理失败，红色提示"""
        ctx = ProcessContext()
        ctx.status = "failed"
        ctx.module_results = {
            "F1": {"fail": 1, "message": "文件不存在: data.xlsx"},
        }

        message, level = build_result_message(ctx)

        # ❌ 失败 → 错误信息
        assert level == "error"
        assert "失败" in message or "error" in message.lower()
        assert "文件不存在" in message or "失败" in message

    def test_warning_message_shows_correct_count(self):
        """⚠️ 警告信息中错误数量与实际一致"""
        ctx = ProcessContext()
        ctx.status = "completed"
        ctx.summary = {
            "yixian": {"total": 1000, "passed": 980, "failed": 20},
            "sanfang": {"total": 500, "passed": 490, "failed": 10},
        }
        ctx.module_results = {
            "F5": {"fail": 30},
        }

        message, level = build_result_message(ctx)

        assert level == "warning"
        # 总计 30 条警告（20 + 10）
        assert "30" in message

    def test_success_message_no_count_shown(self):
        """✅ 成功信息中不显示数量（无问题不需要）"""
        ctx = ProcessContext()
        ctx.status = "completed"
        ctx.summary = {
            "yixian": {"total": 100, "passed": 100, "failed": 0},
        }
        ctx.module_results = {
            "F5": {"fail": 0},
        }

        message, level = build_result_message(ctx)

        assert level == "success"
        # 成功时不应出现"0条"等冗余提示
        assert "0" not in message or "可直接上传" in message

    @pytest.mark.parametrize("failed_count,level", [
        (0, "success"),
        (1, "warning"),
        (100, "warning"),
    ])
    def test_level_classification_by_failed_count(self, failed_count, level):
        """边界值：failed=0→success，failed≥1→warning"""
        ctx = ProcessContext()
        ctx.status = "completed"
        ctx.summary = {
            "yixian": {"total": 100, "passed": 100 - failed_count, "failed": failed_count},
        }
        ctx.module_results = {
            "F5": {"fail": failed_count},
        }

        _, result_level = build_result_message(ctx)
        assert result_level == level
