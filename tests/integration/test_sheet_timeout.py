"""
集成测试 - Sheet超时自动选择测试

测试场景：
- Sheet超时自动选择组件的单元测试

运行方式：
    pytest tests/integration/test_sheet_timeout.py -v
"""

import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

# 确保项目根目录在 Python 路径中
PROJECT_ROOT = Path(__file__).parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


class TestSheetTimeoutComponent:
    """Sheet超时选择组件测试"""

    @pytest.fixture
    def temp_dir(self):
        """创建临时目录"""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    def test_sheet_dialog_creation(self):
        """测试Sheet选择对话框创建"""
        from ui.widgets.sheet_select_dialog import SheetSelectDialog

        dialog = SheetSelectDialog(
            sheet_names=["Sheet1", "Data", "Report"],
            timeout_seconds=5
        )

        assert dialog.sheet_names == ["Sheet1", "Data", "Report"]
        assert dialog.timeout_seconds == 5
        dialog.close()

    def test_sheet_dialog_callback(self):
        """测试Sheet选择回调函数创建"""
        from ui.widgets.sheet_select_dialog import create_sheet_selection_callback

        callback = create_sheet_selection_callback()

        # 单Sheet情况应直接返回
        result = callback(["Sheet1"])
        assert result == "Sheet1"


class TestMultiSheetExcel:
    """多Sheet Excel测试"""

    def test_create_multi_sheet_excel(self, temp_dir):
        """测试创建多Sheet Excel文件"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = Workbook()

        # 移除默认Sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 创建多个Sheet
        ws1 = wb.create_sheet('数据')
        ws1['A1'] = '邮箱'
        ws1['A2'] = 'test@example.com'

        ws2 = wb.create_sheet('备份')
        ws2['A1'] = '邮箱'

        file_path = os.path.join(temp_dir, 'multi_sheet.xlsx')
        wb.save(file_path)

        # 验证文件已创建
        assert os.path.exists(file_path)

        # 验证可以读取
        wb2 = openpyxl.load_workbook(file_path)
        assert '数据' in wb2.sheetnames
        assert '备份' in wb2.sheetnames


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
