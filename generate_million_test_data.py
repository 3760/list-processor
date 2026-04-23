"""
生成百万级测试数据（高效版）
- 55个字段
- 50个字段依据数据字典生成
- 一线/三方/HW各100万行
- 重复率5%
"""

import os
import random
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing

# 尝试导入 xlsxwriter（性能更好）
try:
    import xlsxwriter
    USE_XLSXWRITER = True
except ImportError:
    USE_XLSXWRITER = False
    print("提示: xlsxwriter 未安装，将使用 openpyxl（较慢）")
    print("安装命令: pip install xlsxwriter")

# 尝试导入 openpyxl（用于 CSV 转 Excel）
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 基础路径
BASE_PATH = "/Users/mars/Desktop/00_Work-Ipsos/01_项目/00_售前项目/20260315_华为/名单处理工具-AI流程/业务验收/测试数据/百万级测试数据"
os.makedirs(BASE_PATH, exist_ok=True)

# ========== 1. 定义55个字段规范 ==========
FIELD_SPEC = {
    # === 基础必填字段（5个）===
    "姓名": {"required": True, "type": "text"},
    "邮箱": {"required": True, "type": "email"},
    "手机": {"required": False, "type": "phone"},
    "客户来源": {"required": True, "type": "dict", "dict_id": "A1"},
    "客户等级": {"required": False, "type": "dict", "dict_id": "A2"},

    # === 客户属性字典字段（20个，A3-A22）===
    "客户类型": {"required": True, "type": "dict", "dict_id": "A3"},
    "客户状态": {"required": True, "type": "dict", "dict_id": "A4"},
    "客户行业": {"required": True, "type": "dict", "dict_id": "A5"},
    "客户规模": {"required": False, "type": "dict", "dict_id": "A6"},
    "客户地区": {"required": True, "type": "dict", "dict_id": "A7"},
    "客户性质": {"required": False, "type": "dict", "dict_id": "A8"},
    "客户来源渠道": {"required": True, "type": "dict", "dict_id": "A9"},
    "客户优先级": {"required": False, "type": "dict", "dict_id": "A10"},
    "客户信用等级": {"required": False, "type": "dict", "dict_id": "A11"},
    "客户价值分层": {"required": True, "type": "dict", "dict_id": "A12"},
    "客户生命周期": {"required": True, "type": "dict", "dict_id": "A13"},
    "客户接触方式": {"required": False, "type": "dict", "dict_id": "A14"},
    "客户意向度": {"required": True, "type": "dict", "dict_id": "A15"},
    "客户决策权限": {"required": False, "type": "dict", "dict_id": "A16"},
    "客户跟进阶段": {"required": True, "type": "dict", "dict_id": "A17"},
    "客户满意度": {"required": False, "type": "dict", "dict_id": "A18"},
    "客户风险等级": {"required": False, "type": "dict", "dict_id": "A19"},
    "客户推荐意愿": {"required": True, "type": "dict", "dict_id": "A20"},
    "客户投诉记录": {"required": False, "type": "dict", "dict_id": "A21"},
    "客户黑名单": {"required": False, "type": "dict", "dict_id": "A22"},

    # === 客户偏好字典字段（15个，B1-B15）===
    "产品偏好": {"required": False, "type": "dict", "dict_id": "B1"},
    "服务偏好": {"required": False, "type": "dict", "dict_id": "B2"},
    "沟通偏好": {"required": False, "type": "dict", "dict_id": "B3"},
    "价格敏感度": {"required": True, "type": "dict", "dict_id": "B4"},
    "品牌偏好": {"required": False, "type": "dict", "dict_id": "B5"},
    "购买频次偏好": {"required": False, "type": "dict", "dict_id": "B6"},
    "支付方式偏好": {"required": True, "type": "dict", "dict_id": "B7"},
    "配送方式偏好": {"required": False, "type": "dict", "dict_id": "B8"},
    "促销敏感度": {"required": False, "type": "dict", "dict_id": "B9"},
    "渠道偏好": {"required": True, "type": "dict", "dict_id": "B10"},
    "时间偏好": {"required": False, "type": "dict", "dict_id": "B11"},
    "社交媒体活跃度": {"required": False, "type": "dict", "dict_id": "B12"},
    "内容偏好": {"required": True, "type": "dict", "dict_id": "B13"},
    "活动偏好": {"required": False, "type": "dict", "dict_id": "B14"},
    "会员等级": {"required": False, "type": "dict", "dict_id": "B15"},

    # === 客户行为字典字段（15个，C1-C15）===
    "购买行为": {"required": True, "type": "dict", "dict_id": "C1"},
    "咨询行为": {"required": False, "type": "dict", "dict_id": "C2"},
    "投诉行为": {"required": False, "type": "dict", "dict_id": "C3"},
    "退款行为": {"required": True, "type": "dict", "dict_id": "C4"},
    "复购行为": {"required": False, "type": "dict", "dict_id": "C5"},
    "浏览行为": {"required": False, "type": "dict", "dict_id": "C6"},
    "收藏行为": {"required": True, "type": "dict", "dict_id": "C7"},
    "分享行为": {"required": False, "type": "dict", "dict_id": "C8"},
    "评价行为": {"required": False, "type": "dict", "dict_id": "C9"},
    "签收行为": {"required": True, "type": "dict", "dict_id": "C10"},
    "退货行为": {"required": False, "type": "dict", "dict_id": "C11"},
    "换货行为": {"required": False, "type": "dict", "dict_id": "C12"},
    "使用行为": {"required": True, "type": "dict", "dict_id": "C13"},
    "激活行为": {"required": False, "type": "dict", "dict_id": "C14"},
    "到期行为": {"required": False, "type": "dict", "dict_id": "C15"},

    # === 文本输入字段（5个，无字典）===
    "备注": {"required": False, "type": "text"},
    "详细地址": {"required": False, "type": "text"},
    "公司名称": {"required": False, "type": "text"},
    "职位": {"required": False, "type": "text"},
    "身份证号": {"required": False, "type": "id_card"},

    # === 日期字段（2个）===
    "注册日期": {"required": True, "type": "date"},
    "最近购买日期": {"required": False, "type": "date"},

    # === 数值字段（3个）===
    "累计消费金额": {"required": False, "type": "number"},
    "购买次数": {"required": False, "type": "number"},
    "积分": {"required": False, "type": "number"},
}

print(f"字段总数: {len(FIELD_SPEC)}")

# ========== 2. 定义数据字典（50个，字典间不重复）==========
DATA_DICT = {
    # A1 客户来源 (5个值)
    "A1": ["线上推广", "线下活动", "电话营销", "自然流量", "老客推荐"],
    # A2 客户等级 (5个值)
    "A2": ["VIP客户", "普通客户", "潜在客户", "流失客户", "休眠客户"],
    # A3 客户类型 (5个值)
    "A3": ["个人客户", "企业客户", "政府客户", "教育客户", "医疗客户"],
    # A4 客户状态 (4个值)
    "A4": ["活跃", "沉默", "流失", "新客"],
    # A5 客户行业 (10个值)
    "A5": ["互联网", "金融", "制造业", "零售", "房地产", "教育", "医疗", "物流", "旅游", "餐饮"],
    # A6 客户规模 (5个值)
    "A6": ["微型", "小型", "中型", "大型", "超大型"],
    # A7 客户地区 (8个值)
    "A7": ["华北", "华东", "华南", "华中", "西南", "西北", "东北", "港澳台"],
    # A8 客户性质 (5个值)
    "A8": ["国有企业", "民营企业", "外资企业", "合资企业", "上市公司"],
    # A9 客户来源渠道 (6个值)
    "A9": ["搜索引擎", "社交媒体", "广告投放", "合作伙伴", "线下展会", "电话销售"],
    # A10 客户优先级 (4个值)
    "A10": ["高", "中", "低", "普通"],
    # A11 客户信用等级 (5个值)
    "A11": ["AAA", "AA", "A", "B", "C"],
    # A12 客户价值分层 (4个值)
    "A12": ["高价值", "中价值", "低价值", "负价值"],
    # A13 客户生命周期 (5个值)
    "A13": ["引入期", "成长期", "成熟期", "衰退期", "休眠期"],
    # A14 客户接触方式 (5个值)
    "A14": ["电话", "邮件", "短信", "微信", "上门"],
    # A15 客户意向度 (5个值)
    "A15": ["强意向", "中意向", "弱意向", "无意向", "待定"],
    # A16 客户决策权限 (4个值)
    "A16": ["决策人", "负责人", "参与者", "否决人"],
    # A17 客户跟进阶段 (6个值)
    "A17": ["初步接触", "需求确认", "方案报价", "商务谈判", "合同签订", "成交"],
    # A18 客户满意度 (5个值)
    "A18": ["非常满意", "满意", "一般", "不满意", "非常不满意"],
    # A19 客户风险等级 (4个值)
    "A19": ["低风险", "中风险", "高风险", "极高风险"],
    # A20 客户推荐意愿 (5个值)
    "A20": ["非常愿意", "愿意", "一般", "不愿意", "非常不愿意"],
    # A21 客户投诉记录 (3个值)
    "A21": ["无投诉", "轻度投诉", "重度投诉"],
    # A22 客户黑名单 (2个值)
    "A22": ["是", "否"],

    # B1 产品偏好 (8个值)
    "B1": ["电子产品", "服装鞋帽", "食品饮料", "家居用品", "美妆护肤", "母婴用品", "运动户外", "图书音像"],
    # B2 服务偏好 (6个值)
    "B2": ["售前咨询", "售后服务", "物流配送", "退换货", "会员服务", "技术支持"],
    # B3 沟通偏好 (5个值)
    "B3": ["电话沟通", "文字交流", "视频会议", "上门服务", "自助服务"],
    # B4 价格敏感度 (4个值)
    "B4": ["极度敏感", "敏感", "一般", "不敏感"],
    # B5 品牌偏好 (8个值)
    "B5": ["苹果", "华为", "小米", "三星", "OPPO", "vivo", "荣耀", "一加"],
    # B6 购买频次偏好 (5个值)
    "B6": ["每天", "每周", "每月", "每季度", "每年"],
    # B7 支付方式偏好 (6个值)
    "B7": ["微信支付", "支付宝", "银行卡", "现金", "分期付款", "信用支付"],
    # B8 配送方式偏好 (5个值)
    "B8": ["快递", "自提", "送货上门", "当日达", "次日达"],
    # B9 促销敏感度 (4个值)
    "B9": ["极度敏感", "敏感", "一般", "不敏感"],
    # B10 渠道偏好 (7个值)
    "B10": ["官网", "APP", "小程序", "实体店", "第三方平台", "直播", "社群"],
    # B11 时间偏好 (4个值)
    "B11": ["工作日", "周末", "节假日", "随时"],
    # B12 社交媒体活跃度 (5个值)
    "B12": ["非常活跃", "活跃", "一般", "不活跃", "潜水"],
    # B13 内容偏好 (8个值)
    "B13": ["新闻资讯", "娱乐八卦", "生活方式", "科技数码", "财经商业", "体育运动", "教育培训", "健康养生"],
    # B14 活动偏好 (6个值)
    "B14": ["促销活动", "新品体验", "会员活动", "线下沙龙", "抽奖活动", "公益活动"],
    # B15 会员等级 (5个值)
    "B15": ["普通会员", "银卡会员", "金卡会员", "白金会员", "钻石会员"],

    # C1 购买行为 (5个值)
    "C1": ["线上购买", "线下购买", "线上线下", "从未购买", "预购"],
    # C2 咨询行为 (5个值)
    "C2": ["频繁咨询", "偶尔咨询", "从不咨询", "咨询后购买", "咨询后退款"],
    # C3 投诉行为 (4个值)
    "C3": ["无投诉", "有投诉已解决", "有投诉未解决", "重复投诉"],
    # C4 退款行为 (5个值)
    "C4": ["无退款", "偶尔退款", "经常退款", "退款后复购", "仅退款"],
    # C5 复购行为 (5个值)
    "C5": ["持续复购", "偶尔复购", "一次购买", "复购周期长", "复购间隔短"],
    # C6 浏览行为 (5个值)
    "C6": ["深度浏览", "浅度浏览", "浏览后购买", "浏览后离开", "持续关注"],
    # C7 收藏行为 (4个值)
    "C7": ["经常收藏", "偶尔收藏", "收藏后购买", "从不收藏"],
    # C8 分享行为 (5个值)
    "C8": ["经常分享", "偶尔分享", "分享后引流", "被动分享", "从不分享"],
    # C9 评价行为 (5个值)
    "C9": ["积极好评", "消极差评", "默认好评", "从不评价", "追加评价"],
    # C10 签收行为 (4个值)
    "C10": ["正常签收", "延迟签收", "拒收", "代收"],
    # C11 退货行为 (5个值)
    "C11": ["无退货", "偶尔退货", "经常退货", "退货后换货", "仅退货"],
    # C12 换货行为 (4个值)
    "C12": ["尺码换货", "颜色换货", "型号换货", "无换货"],
    # C13 使用行为 (5个值)
    "C13": ["正常使用", "高频使用", "低频使用", "偶尔使用", "从未使用"],
    # C14 激活行为 (4个值)
    "C14": ["已激活", "未激活", "激活过期", "激活异常"],
    # C15 到期行为 (5个值)
    "C15": ["正常到期", "提前续费", "到期流失", "到期提醒", "到期未处理"],
}

print(f"数据字典总数: {len(DATA_DICT)}")

# ========== 3. 预生成数据池 ==========
# 预先生成常用数据池，避免每次随机生成
random.seed(42)
SURENAMES = ["张", "王", "李", "赵", "刘", "陈", "杨", "黄", "周", "吴", "徐", "孙", "胡", "朱", "高", "林", "何", "郭", "马", "罗", "梁", "宋", "郑", "谢", "韩", "唐", "冯", "于", "董", "萧", "程", "曹", "袁", "邓", "许", "傅", "沈", "曾", "彭", "吕"]
NAMES = ["伟", "芳", "娜", "敏", "静", "丽", "强", "磊", "军", "洋", "勇", "艳", "杰", "娟", "涛", "明", "超", "秀英", "桂英", "丹", "华", "红", "玉兰", "玉珍", "凤英", "玲", "桂兰"]
DOMAINS = ["qq.com", "163.com", "126.com", "gmail.com", "sina.com", "outlook.com", "hotmail.com"]
PHONE_PREFIXES = ["130", "131", "132", "133", "134", "135", "136", "137", "138", "139", "150", "151", "152", "153", "155", "156", "157", "158", "159", "170", "171", "173", "175", "176", "177", "178", "180", "181", "182", "183", "184", "185", "186", "187", "188", "189"]
CITIES = ["北京市朝阳区", "上海市浦东新区", "广州市天河区", "深圳市南山区", "杭州市西湖区", "成都市高新区", "南京市鼓楼区", "武汉市洪山区"]
COMPANY_PREFIXES = ["腾讯", "阿里", "百度", "字节", "美团", "京东", "华为", "小米", "联想", "海尔", "格力", "比亚迪", "蔚来", "小鹏", "理想"]
POSITIONS = ["经理", "总监", "主管", "专员", "助理", "工程师", "顾问", "董事", "总经理", "CEO"]

# 预生成数据池
NAME_POOL = [random.choice(SURENAMES) + random.choice(NAMES) for _ in range(10000)]
EMAIL_POOL = [f"{n}{random.randint(1,9999)}@{random.choice(DOMAINS)}" for n in NAME_POOL[:5000]]
PHONE_POOL = [random.choice(PHONE_PREFIXES) + "".join([str(random.randint(0, 9)) for _ in range(8)]) for _ in range(10000)]
ADDRESS_POOL = [random.choice(CITIES) + f"XX路{random.randint(1,999)}号" for _ in range(1000)]
COMPANY_POOL = [random.choice(COMPANY_PREFIXES) + "科技有限公司" for _ in range(1000)]
POSITION_POOL = POSITIONS.copy()

# 日期池
START_DATE = datetime(2020, 1, 1)
END_DATE = datetime(2024, 12, 31)
DATE_POOL = [(START_DATE + timedelta(days=random.randint(0, (END_DATE - START_DATE).days))).strftime("%Y-%m-%d") for _ in range(5000)]

# 数值池
AMOUNT_POOL = [random.randint(0, 1000000) for _ in range(1000)]
COUNT_POOL = [random.randint(0, 1000) for _ in range(1000)]
SCORE_POOL = [random.randint(0, 100000) for _ in range(1000)]

def get_name():
    return random.choice(NAME_POOL)

def get_email():
    return random.choice(EMAIL_POOL)

def get_phone():
    return random.choice(PHONE_POOL)

def get_address():
    return random.choice(ADDRESS_POOL)

def get_company():
    return random.choice(COMPANY_POOL)

def get_position():
    return random.choice(POSITION_POOL)

def get_date():
    return random.choice(DATE_POOL)

def get_amount():
    return random.choice(AMOUNT_POOL)

def get_count():
    return random.choice(COUNT_POOL)

def get_score():
    return random.choice(SCORE_POOL)

def get_dict_value(dict_id):
    return random.choice(DATA_DICT.get(dict_id, [""]))

# ========== 4. 生成单条记录 ==========
def generate_record():
    """生成单条记录"""
    record = {}
    name = get_name()

    for field_name, spec in FIELD_SPEC.items():
        field_type = spec["type"]

        if field_type == "text":
            if field_name == "姓名":
                record[field_name] = name
            elif field_name == "邮箱":
                record[field_name] = get_email()
            elif field_name == "手机":
                record[field_name] = get_phone()
            elif field_name == "备注":
                record[field_name] = f"备注{random.randint(1,1000)}"
            elif field_name == "详细地址":
                record[field_name] = get_address()
            elif field_name == "公司名称":
                record[field_name] = get_company()
            elif field_name == "职位":
                record[field_name] = get_position()
            elif field_name == "身份证号":
                record[field_name] = ""

        elif field_type == "date":
            record[field_name] = get_date()

        elif field_type == "number":
            if field_name == "累计消费金额":
                record[field_name] = get_amount()
            elif field_name == "购买次数":
                record[field_name] = get_count()
            else:
                record[field_name] = get_score()

        elif field_type == "dict":
            record[field_name] = get_dict_value(spec["dict_id"])

        # 必填字段处理
        if spec.get("required") and (field_name not in record or not record[field_name]):
            if field_type == "text":
                record[field_name] = "未填写"

    return record

# ========== 5. 批量生成函数 ==========
def generate_batch(args):
    """批量生成记录"""
    n_rows, unique_ratio, batch_id, total_batches = args

    random.seed(42 + batch_id)
    n_unique = int(n_rows * unique_ratio)
    n_duplicate = n_rows - n_unique

    records = []
    for _ in range(n_unique):
        records.append(generate_record())

    # 添加重复记录
    for _ in range(n_duplicate):
        records.append(random.choice(records[:n_unique]))

    return records

# ========== 6. 生成数据集 ==========
def generate_dataset(n_rows=1000000, filename="test.xlsx", n_batches=100):
    """生成数据集"""
    print(f"\n开始生成 {filename}，共 {n_rows:,} 行...")

    batch_size = n_rows // n_batches
    unique_ratio = 0.95

    all_records = []
    for i in range(n_batches):
        if i == n_batches - 1:
            batch_rows = n_rows - len(all_records)
        else:
            batch_rows = batch_size

        args = (batch_rows, unique_ratio, i, n_batches)
        batch_records = generate_batch(args)
        all_records.extend(batch_records)

        if (i + 1) % 10 == 0:
            print(f"  进度: {(i+1)*100//n_batches}% ({len(all_records):,} 行)")

    # 打乱顺序
    random.shuffle(all_records)

    print(f"  创建DataFrame...")
    df = pd.DataFrame(all_records)
    
    # 获取字段名
    columns = list(df.columns)

    # 保存
    output_path = os.path.join(BASE_PATH, filename)
    print(f"  保存到: {output_path}")
    print(f"  文件大小约: {len(df) * len(df.columns) * 50 / 1024 / 1024:.1f} MB")

    # 统一使用 openpyxl write_only 模式（流式写入，速度快，保存不卡顿）
    _save_with_openpyxl(df, output_path)

    # 验证
    total_rows = len(df)
    unique_rows = df.drop_duplicates().shape[0]
    dup_rate = (total_rows - unique_rows) / total_rows * 100
    print(f"  验证: 总行数={total_rows:,}, 唯一行={unique_rows:,}, 重复率={dup_rate:.2f}%")

    return df


def _csv_to_xlsx(csv_path, output_path):
    """使用 xlsxwriter 将 CSV 转换为 Excel（内存友好）"""
    print("    Step 2: xlsxwriter 转换 CSV → Excel...")
    
    wb = xlsxwriter.Workbook(output_path, {
        'constant_memory': True,
        'use_zip64': True  # 支持大于4GB的文件
    })
    ws = wb.add_worksheet('Data')
    
    # 使用 xlsxwriter 的 csv 读取方式
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        import csv as csv_module
        reader = csv_module.reader(f)
        for row_idx, row in enumerate(reader):
            for col_idx, value in enumerate(row):
                ws.write(row_idx, col_idx, value)
            if row_idx > 0 and row_idx % 100000 == 0:
                print(f"      转换进度: {row_idx:,} 行已写入")
    
    print("    Step 3: 关闭文件...")
    wb.close()


def _save_with_xlsxwriter(df, output_path, columns):
    """高效写入：CSV → xlsxwriter → Excel"""
    import time
    start_time = time.time()
    
    print("  使用 CSV 中转 → xlsxwriter → Excel 方案...")
    
    # Step 1: 先写入 CSV（极快）
    csv_path = output_path.replace('.xlsx', '_temp.csv')
    print("    Step 1: 写入 CSV...")
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    csv_size = os.path.getsize(csv_path) / 1024 / 1024
    print(f"    CSV 写入完成! 大小: {csv_size:.1f} MB")
    
    # 释放 DataFrame 内存
    del df
    
    # Step 2: xlsxwriter 转换
    _csv_to_xlsx(csv_path, output_path)
    
    # 删除临时 CSV
    os.remove(csv_path)
    
    elapsed = time.time() - start_time
    print(f"  保存完成! 耗时: {elapsed:.1f} 秒")


def _csv_to_openpyxl(csv_path, output_path):
    """使用 openpyxl write_only 模式将 CSV 转换为 Excel（流式写入）"""
    print("    Step 2: openpyxl write_only 模式流式写入...")
    
    # write_only 模式：流式写入，专门处理大数据
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet('Data')
    
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        import csv as csv_module
        reader = csv_module.reader(f)
        for row_idx, row in enumerate(reader):
            ws.append(row)  # 流式追加，非常快
            if row_idx > 0 and row_idx % 100000 == 0:
                print(f"      写入进度: {row_idx:,} 行")
    
    print("    Step 3: 保存文件（流式写入，保存很快）...")
    wb.save(output_path)
    wb.close()
    print("    保存完成！")


def _save_with_openpyxl(df, output_path):
    """写入：CSV → openpyxl → Excel"""
    import time
    start_time = time.time()
    
    print("  使用 CSV 中转 → openpyxl → Excel 方案...")
    
    # Step 1: 先写入 CSV
    csv_path = output_path.replace('.xlsx', '_temp.csv')
    print("    Step 1: 写入 CSV...")
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    csv_size = os.path.getsize(csv_path) / 1024 / 1024
    print(f"    CSV 写入完成! 大小: {csv_size:.1f} MB")
    
    # 释放 DataFrame 内存
    del df
    
    # Step 2: openpyxl 转换
    _csv_to_openpyxl(csv_path, output_path)
    
    # 删除临时 CSV
    os.remove(csv_path)
    
    elapsed = time.time() - start_time
    print(f"  保存完成! 耗时: {elapsed:.1f} 秒")

# ========== 7. 生成字段规范文件 ==========
def generate_field_spec_file():
    """生成字段规范Excel文件"""
    rows = []
    for i, (field_name, spec) in enumerate(FIELD_SPEC.items(), 1):
        dict_id = spec.get("dict_id", "")
        required = "是" if spec.get("required") else "否"
        field_type = spec.get("type", "文本型")

        rows.append({
            "序号": i,
            "属性名称": field_name,
            "属性code": field_name,
            "属性类型": "客户属性",
            "数据类型": field_type,
            "数据子类型": "输入框" if not dict_id else "下拉单选",
            "长度上限": "",
            "数据字典": dict_id,
            "属性值必填": required,
            "验证规则": ""
        })

    df = pd.DataFrame(rows)
    output_path = os.path.join(BASE_PATH, "field_spec_百万级.xlsx")
    df.to_excel(output_path, index=False, engine='openpyxl')
    print(f"字段规范文件已生成: {output_path}")

# ========== 8. 生成数据字典文件 ==========
def generate_data_dict_file():
    """
    生成数据字典Excel文件（参考 租户字典导入模版.xlsx 格式）。
    
    格式：每个字典占 2 列（Code、Details），列间用空列分隔
    - 第1行：dict_id（A1、A2...）
    - 第2行：字典名称（客户来源、客户等级...）
    - 第3行：Code、Details 表头
    - 第4行起：数据
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # dict_id 到中文名的映射
    dict_name_map = {
        "A1": "客户来源", "A2": "客户等级", "A3": "客户类型", "A4": "客户状态",
        "A5": "客户行业", "A6": "客户规模", "A7": "客户地区", "A8": "客户性质",
        "A9": "客户来源渠道", "A10": "客户优先级", "A11": "客户信用等级", "A12": "客户价值分层",
        "A13": "客户生命周期", "A14": "客户接触方式", "A15": "客户意向度", "A16": "客户决策权限",
        "A17": "客户跟进阶段", "A18": "客户满意度", "A19": "客户风险等级", "A20": "客户推荐意愿",
        "A21": "客户投诉记录", "A22": "客户黑名单",
        "B1": "产品偏好", "B2": "服务偏好", "B3": "沟通偏好", "B4": "价格敏感度",
        "B5": "品牌偏好", "B6": "购买频次偏好", "B7": "支付方式偏好", "B8": "配送方式偏好",
        "B9": "促销敏感度", "B10": "渠道偏好", "B11": "时间偏好", "B12": "社交媒体活跃度",
        "B13": "内容偏好", "B14": "活动偏好", "B15": "会员等级",
        "C1": "购买行为", "C2": "咨询行为", "C3": "投诉行为", "C4": "退款行为",
        "C5": "复购行为", "C6": "浏览行为", "C7": "收藏行为", "C8": "分享行为",
        "C9": "评价行为", "C10": "签收行为", "C11": "退货行为", "C12": "换货行为",
        "C13": "使用行为", "C14": "激活行为", "C15": "到期行为",
    }
    
    # code 前缀映射
    code_map = {
        "A1": "SRC", "A2": "LVL", "A3": "TYPE", "A4": "STAT",
        "A5": "IND", "A6": "SIZE", "A7": "REG", "A8": "NATURE",
        "A9": "CHNL", "A10": "PRI", "A11": "CRED", "A12": "VAL",
        "A13": "LC", "A14": "CON", "A15": "INT", "A16": "AUTH",
        "A17": "STAGE", "A18": "SAT", "A19": "RISK", "A20": "REC",
        "A21": "CMP", "A22": "BLK",
        "B1": "PROD", "B2": "SVC", "B3": "COMM", "B4": "PRICE",
        "B5": "BRAND", "B6": "FREQ", "B7": "PAY", "B8": "DLV",
        "B9": "PROM", "B10": "CHAN", "B11": "TIME", "B12": "SOC",
        "B13": "CONT", "B14": "EVT", "B15": "MEM",
        "C1": "BUY", "C2": "INQ", "C3": "CMP", "C4": "REF",
        "C5": "REP", "C6": "BRW", "C7": "FAV", "C8": "SHR",
        "C9": "REV", "C10": "SGN", "C11": "RET", "C12": "EXC",
        "C13": "USE", "C14": "ACT", "C15": "EXP"
    }
    
    # 按顺序遍历每个字典，写入 2 列（Code、Details），列间用空列分隔
    col_idx = 0
    for dict_id in sorted(DATA_DICT.keys()):
        dict_name = dict_name_map.get(dict_id, dict_id)
        values = DATA_DICT[dict_id]
        prefix = code_map.get(dict_id, dict_id)
        
        # 第1行：dict_id
        ws.cell(row=1, column=col_idx + 1, value=dict_id)
        
        # 第2行：字典名称
        ws.cell(row=2, column=col_idx + 1, value=dict_name)
        
        # 第3行：Code、Details 表头
        ws.cell(row=3, column=col_idx + 1, value="Code")
        ws.cell(row=3, column=col_idx + 2, value="Details")
        
        # 数据行（第4行起）
        for row_offset, detail in enumerate(values):
            code = f"{prefix}_{row_offset + 1:02d}"
            ws.cell(row=4 + row_offset, column=col_idx + 1, value=code)
            ws.cell(row=4 + row_offset, column=col_idx + 2, value=detail)
        
        col_idx += 2  # 每个字典占 2 列（Code、Details）
    
    output_path = os.path.join(BASE_PATH, "data_dict_百万级.xlsx")
    wb.save(output_path)
    print(f"数据字典文件已生成: {output_path}")

# ========== 主程序 ==========
if __name__ == "__main__":
    print("=" * 60)
    print("开始生成百万级测试数据")
    print("=" * 60)

    # 生成字段规范和数据字典
    print("\n[1/5] 生成字段规范文件...")
    generate_field_spec_file()

    print("\n[2/5] 生成数据字典文件...")
    generate_data_dict_file()

    # 生成一线名单
    print("\n[3/5] 生成一线名单...")
    #generate_dataset(n_rows=200000, filename="一线名单_20万.xlsx")

    # 生成三方名单
    print("\n[4/5] 生成三方名单...")
    #generate_dataset(n_rows=200000, filename="三方名单_20万.xlsx")

    # 生成HW名单
    print("\n[5/5] 生成HW名单...")
    #generate_dataset(n_rows=200000, filename="HW名单_20万.xlsx")

    print("\n" + "=" * 60)
    print("所有测试数据生成完成！")
    print(f"输出目录: {BASE_PATH}")
    print("=" * 60)
